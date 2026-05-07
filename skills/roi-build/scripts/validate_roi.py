#!/usr/bin/env python3
"""
Inline validator (Stage A) for ROI workbooks.

Usage:
    python validate_roi.py <workbook.xlsx> [--user-high-stakes]

Output: JSON with red_flags, warnings, passed_checks, trigger_review decision.

Rules: see references/sanity_check_rules.md
"""

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


# Labels in 5_Output whose B/C/D cells must be formulas (load-bearing outputs).
# If hardcoded, future input changes won't propagate — silent staleness risk.
# Match is substring + case-insensitive to tolerate label variations across feature workbooks.
FORMULA_REQUIRED_LABELS = (
    'Total Value (THB)',
    'Y1 ROI (Value ÷ Cost)',
    '3-Year Total Value',
    '3-Year ROI',
)

CROSS_SHEET_REF_RE = re.compile(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)")
UNRESOLVED_REF_RE = re.compile(r'(?<![A-Za-z])ref:[A-Za-z0-9_\-]+')


def _check_formula_integrity(wb_f):
    """FI-01..FI-03: load-bearing cells must be formulas, cross-sheet refs must resolve,
    and no '=ref:' build-time placeholders may remain.

    Reads the workbook with data_only=False so we see formula strings (starting with '='),
    not their cached computed values.

    Returns dict with red_flags + passed_checks lists.
    """
    red_flags = []
    passed_checks = []

    if wb_f is None:
        return {'red_flags': red_flags, 'passed_checks': passed_checks}

    # === FI-01: Load-bearing 5_Output cells must be formulas ===
    try:
        ws_out = wb_f['5_Output']
        hardcoded = []
        for row in ws_out.iter_rows(values_only=False):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                label = cell.value.strip()
                matched = next(
                    (lbl for lbl in FORMULA_REQUIRED_LABELS if lbl.lower() in label.lower()),
                    None,
                )
                if not matched:
                    continue
                # Check B, C, D in same row
                for col in (2, 3, 4):
                    target = ws_out.cell(row=cell.row, column=col)
                    val = target.value
                    if val is None:
                        continue
                    is_formula = isinstance(val, str) and val.startswith('=')
                    if not is_formula:
                        snippet = val if not isinstance(val, str) else val[:30]
                        hardcoded.append(f"{target.coordinate} [{matched}] = {snippet!r}")
        if hardcoded:
            red_flags.append({
                'rule': 'FI-01',
                'name': 'load_bearing_cell_hardcoded',
                'value': len(hardcoded),
                'severity': 'red',
                'message': (
                    f"{len(hardcoded)} load-bearing cell(s) hardcoded instead of formula: "
                    + '; '.join(hardcoded[:3])
                    + ('...' if len(hardcoded) > 3 else '')
                    + '. Future input changes will NOT propagate — restore formula chain.'
                ),
            })
        else:
            passed_checks.append('load_bearing_cells_are_formulas')
    except KeyError:
        pass

    # === FI-02: Cross-sheet refs in 5_Output must resolve to a non-empty target ===
    try:
        ws_out = wb_f['5_Output']
        broken = []
        for row in ws_out.iter_rows(values_only=False):
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                for m in CROSS_SHEET_REF_RE.finditer(cell.value):
                    ref_sheet, ref_col, ref_row = m.group(1), m.group(2), m.group(3)
                    if ref_sheet not in wb_f.sheetnames:
                        broken.append(f"{cell.coordinate} → '{ref_sheet}' (sheet missing)")
                        continue
                    target_val = wb_f[ref_sheet][f'{ref_col}{ref_row}'].value
                    if target_val is None:
                        broken.append(
                            f"{cell.coordinate} → '{ref_sheet}'!{ref_col}{ref_row} (target empty)"
                        )
        if broken:
            red_flags.append({
                'rule': 'FI-02',
                'name': 'cross_sheet_ref_broken',
                'value': len(broken),
                'severity': 'red',
                'message': (
                    f"{len(broken)} broken cross-sheet ref(s) in 5_Output: "
                    + '; '.join(broken[:3])
                    + ('...' if len(broken) > 3 else '')
                ),
            })
        else:
            passed_checks.append('cross_sheet_refs_resolve')
    except KeyError:
        pass

    # === FI-03: Unresolved 'ref:' placeholders from build_roi_workbook resolver ===
    try:
        unresolved = []
        for sheet_name in wb_f.sheetnames:
            ws = wb_f[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    if UNRESOLVED_REF_RE.search(cell.value):
                        snippet = cell.value[:40]
                        unresolved.append(f"{sheet_name}!{cell.coordinate}: {snippet!r}")
        if unresolved:
            red_flags.append({
                'rule': 'FI-03',
                'name': 'unresolved_ref_placeholder',
                'value': len(unresolved),
                'severity': 'red',
                'message': (
                    f"{len(unresolved)} unresolved 'ref:' placeholder(s) — "
                    "build_roi_workbook resolver failed: "
                    + '; '.join(unresolved[:3])
                    + ('...' if len(unresolved) > 3 else '')
                ),
            })
        else:
            passed_checks.append('no_unresolved_ref_placeholders')
    except Exception:
        pass

    return {'red_flags': red_flags, 'passed_checks': passed_checks}


def validate(workbook_path, user_high_stakes=False):
    """Run all sanity rules. Return JSON dict."""
    wb = load_workbook(workbook_path, data_only=True)
    try:
        wb_f = load_workbook(workbook_path, data_only=False)
    except Exception:
        wb_f = None  # FI-* checks gracefully skip if formula view can't load

    red_flags = []
    warnings = []
    passed_checks = []

    # === Read key values from output sheet ===
    try:
        ws_out = wb['5_Output']
    except KeyError:
        return {
            "status": "error",
            "message": "Sheet 5_Output not found — workbook structure invalid",
            "red_flags": [],
            "warnings": [],
            "trigger_review": True,
        }

    # Find ROI cells (Y1 Base, Best, 3yr Base)
    y1_roi_base = y1_roi_best = y3_roi_base = None
    y1_val_base = y1_val_best = None
    for row in ws_out.iter_rows(values_only=False):
        for cell in row:
            if cell.value == 'Y1 ROI (Value ÷ Cost)':
                y1_roi_base = ws_out.cell(row=cell.row, column=3).value
                y1_roi_best = ws_out.cell(row=cell.row, column=4).value
            elif cell.value == '3-Year ROI':
                y3_roi_base = ws_out.cell(row=cell.row, column=3).value
            elif cell.value == 'Total Value (THB)':
                y1_val_base = ws_out.cell(row=cell.row, column=3).value
                y1_val_best = ws_out.cell(row=cell.row, column=4).value

    # === Apply Red Flag Rules ===

    # RF-01: Y1 Best ROI > 20x
    if y1_roi_best is not None and isinstance(y1_roi_best, (int, float)) and y1_roi_best > 20:
        red_flags.append({
            "rule": "RF-01",
            "name": "y1_best_roi_too_high",
            "value": round(y1_roi_best, 2),
            "threshold": 20.0,
            "severity": "red",
            "message": f"Y1 Best ROI {y1_roi_best:.2f}x exceeds 20x cap — likely TAM error or aggressive CFs"
        })

    # RF-02: Y1 Base ROI > 10x
    if y1_roi_base is not None and isinstance(y1_roi_base, (int, float)) and y1_roi_base > 10:
        red_flags.append({
            "rule": "RF-02",
            "name": "y1_base_roi_too_high",
            "value": round(y1_roi_base, 2),
            "threshold": 10.0,
            "severity": "red",
            "message": f"Y1 Base ROI {y1_roi_base:.2f}x is high — verify TAM logic + double-counting"
        })

    # RF-03: TAM logic — read from Inputs
    try:
        ws_inp = wb['2_Inputs']
        for row in ws_inp.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'TAM' in cell.value and ':' in cell.value:
                    tam_val = ws_inp.cell(row=cell.row, column=3).value  # Base
                    if isinstance(tam_val, (int, float)) and tam_val > 1_950_000:  # 1.5x of 1.3M
                        red_flags.append({
                            "rule": "RF-03",
                            "name": "tam_exceeds_org_total",
                            "value": int(tam_val),
                            "threshold": 1_950_000,
                            "severity": "red",
                            "message": f"TAM {tam_val:,.0f} exceeds 1.5× organizational VN — scope likely wrong"
                        })
                    else:
                        passed_checks.append("tam_within_org_total")
                    break
    except KeyError:
        pass

    # RF-06: Formula errors (already checked by recalc.py, but double-check)
    error_strings = ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A']
    found_errors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val and isinstance(val, str) and val in error_strings:
                    found_errors.append(f"{sheet_name}: {val}")
    if found_errors:
        red_flags.append({
            "rule": "RF-06",
            "name": "formula_errors_present",
            "value": len(found_errors),
            "threshold": 0,
            "severity": "red",
            "message": f"Found {len(found_errors)} formula error(s) — fix before review"
        })
    else:
        passed_checks.append("formula_errors_zero")

    # === Apply Warning Rules ===

    # W-01: Y1 Base ROI > 5x
    if y1_roi_base is not None and isinstance(y1_roi_base, (int, float)) and 5 < y1_roi_base <= 10:
        warnings.append({
            "rule": "W-01",
            "name": "y1_roi_above_warning",
            "value": round(y1_roi_base, 2),
            "threshold": 5.0,
            "severity": "warning",
            "message": f"Y1 Base ROI {y1_roi_base:.2f}x exceeds 5x — worth a second look"
        })

    # W-02: T4 + high Pure ROI
    try:
        ws_inp = wb['2_Inputs']
        confidence_tier = None
        for row in ws_inp.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Confidence Tier' in cell.value:
                    # Tier value is in column C of same row
                    tier_val = ws_inp.cell(row=cell.row, column=3).value
                    if isinstance(tier_val, str) and tier_val.strip() in ('T1', 'T2', 'T3', 'T4'):
                        confidence_tier = tier_val.strip()
                    break
        if confidence_tier == 'T4' and y1_roi_base and y1_roi_base > 3:
            warnings.append({
                "rule": "W-02",
                "name": "low_evidence_high_claim",
                "value": f"T4 + {y1_roi_base:.2f}x",
                "severity": "warning",
                "message": "Confidence T4 (low evidence) but Pure Y1 ROI > 3x — risky claim"
            })
    except (KeyError, Exception):
        pass

    # W-04: Best:Worst ratio > 50x
    try:
        if (y1_val_base and y1_val_best
                and isinstance(y1_val_base, (int, float))
                and isinstance(y1_val_best, (int, float))):
            ws_out_data = wb['5_Output']
            for row in ws_out_data.iter_rows(values_only=False):
                for cell in row:
                    if cell.value == 'Total Value (THB)':
                        y1_val_worst = ws_out_data.cell(row=cell.row, column=2).value
                        if (y1_val_worst and isinstance(y1_val_worst, (int, float))
                                and y1_val_worst > 0):
                            ratio = y1_val_best / y1_val_worst
                            if ratio > 50:
                                warnings.append({
                                    "rule": "W-04",
                                    "name": "best_worst_ratio_too_wide",
                                    "value": round(ratio, 1),
                                    "threshold": 50,
                                    "severity": "warning",
                                    "message": f"Best:Worst ratio {ratio:.1f}x indicates high uncertainty"
                                })
                        break
    except Exception:
        pass

    # === RF-04 (strengthened): Cost avoidance separation check ===
    # Old: only checked ratio. New: also verify Hospital ROI breakdown exists
    try:
        ws_vc = wb['3_Value_Calc']
        cost_avoidance_value = revenue_value = ops_value = None
        for row in ws_vc.iter_rows(values_only=False):
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                if 'cost avoidance' in cell.value.lower() and 'value' in cell.value.lower():
                    v = ws_vc.cell(row=cell.row, column=3).value
                    if isinstance(v, (int, float)):
                        cost_avoidance_value = v
                if 'revenue' in cell.value.lower() and 'incremental' in cell.value.lower() and 'THB' in cell.value:
                    v = ws_vc.cell(row=cell.row, column=3).value
                    if isinstance(v, (int, float)):
                        revenue_value = v
                if 'ops savings' in cell.value.lower() and 'THB' in cell.value:
                    v = ws_vc.cell(row=cell.row, column=3).value
                    if isinstance(v, (int, float)):
                        ops_value = v
            # Only need first occurrence (Y1)
            if cost_avoidance_value or revenue_value:
                break

        if cost_avoidance_value:
            total = (cost_avoidance_value or 0) + (revenue_value or 0) + (ops_value or 0)
            if total > 0:
                ca_share = cost_avoidance_value / total

                # Check if separate Hospital ROI breakdown exists
                # 1. Sheet named 'Hospital_ROI' or similar
                # 2. Or Sheet 5 has separate "Hospital ROI" + "System ROI" rows
                has_hospital_sheet = any(
                    'hospital' in s.lower() and 'roi' in s.lower()
                    for s in wb.sheetnames
                )
                has_hospital_section = False
                try:
                    ws_out = wb['5_Output']
                    for row in ws_out.iter_rows(values_only=True):
                        for val in row:
                            if val and isinstance(val, str):
                                vlow = val.lower()
                                if 'hospital roi' in vlow or 'hospital-only' in vlow:
                                    has_hospital_section = True
                                    break
                        if has_hospital_section:
                            break
                except Exception:
                    pass
                has_breakdown = has_hospital_sheet or has_hospital_section

                # Old rule: ratio > 50% → red flag
                if ca_share > 0.5:
                    red_flags.append({
                        "rule": "RF-04",
                        "name": "cost_avoidance_dominant",
                        "value": round(ca_share, 2),
                        "threshold": 0.5,
                        "severity": "red",
                        "message": f"Cost avoidance {ca_share:.0%} of value — split Hospital ROI vs System ROI"
                    })
                # NEW: ratio > 30% but no breakdown → red flag (strengthened)
                elif ca_share > 0.3 and not has_breakdown:
                    red_flags.append({
                        "rule": "RF-04",
                        "name": "cost_avoidance_no_breakdown",
                        "value": round(ca_share, 2),
                        "threshold": 0.3,
                        "severity": "red",
                        "message": f"Cost avoidance {ca_share:.0%} of value but no separate 'Hospital ROI' sheet/section — add explicit breakdown"
                    })
                elif has_breakdown:
                    passed_checks.append("cost_avoidance_separated")
                else:
                    passed_checks.append("cost_avoidance_within_threshold")
    except Exception:
        pass

    # === RF-08: TAM unit mismatch check ===
    # Detect when TAM uses VN/visits but problem describes patient cohort
    try:
        ws_inp = wb['2_Inputs']
        tam_unit = None
        tam_label = None
        for row in ws_inp.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('TAM'):
                    tam_label = cell.value
                    # Unit is in column 5 of same row
                    unit_cell = ws_inp.cell(row=cell.row, column=5).value
                    if unit_cell:
                        tam_unit = str(unit_cell)
                    break
            if tam_unit:
                break

        # Read problem from Sheet 1
        problem_text = ""
        try:
            ws_info = wb['1_Feature_Info']
            for row in ws_info.iter_rows(values_only=False):
                for cell in row:
                    if cell.value == 'Problem (Why)':
                        problem_text = str(ws_info.cell(row=cell.row, column=2).value or "")
                        break
                if problem_text:
                    break
        except Exception:
            pass

        if tam_unit and problem_text:
            tam_is_visits = any(kw in tam_unit.lower() for kw in ['vn', 'visit', 'visits'])
            problem_describes_patients = any(
                kw in problem_text.lower()
                for kw in ['patient', 'ผู้ป่วย', 'cohort', 'user', 'NCD', 'ncd', 'คนไข้', 'unique']
            )
            if tam_is_visits and problem_describes_patients:
                red_flags.append({
                    "rule": "RF-08",
                    "name": "tam_unit_mismatch",
                    "value": f"TAM unit='{tam_unit}', problem mentions patient cohort",
                    "severity": "red",
                    "message": (
                        f"TAM uses '{tam_unit}' but problem describes patient cohort — "
                        "verify unit. 1 NCD patient ≠ 6 NCD visits/year. Risk of 6× double-counting."
                    )
                })
            else:
                passed_checks.append("tam_unit_consistent")
    except Exception:
        pass

    # Confidence tier present check (metadata only, never multiplied)
    try:
        ws_inp = wb['2_Inputs']
        tier_found = False
        for row in ws_inp.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Confidence Tier' in cell.value:
                    tier_val = ws_inp.cell(row=cell.row, column=3).value
                    if isinstance(tier_val, str) and tier_val.strip() in ('T1', 'T2', 'T3', 'T4'):
                        tier_found = True
                    break
        if tier_found:
            passed_checks.append("confidence_tier_present")
    except Exception:
        pass

    # Strategic Fit within cap (Priority Score input integrity)
    try:
        ws_inp = wb['2_Inputs']
        sf_val = None
        for row in ws_inp.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Strategic Fit' in cell.value and 'Priority Score' in cell.value:
                    sf_val = ws_inp.cell(row=cell.row, column=3).value
                    break
        if isinstance(sf_val, (int, float)):
            if sf_val > 1.5:
                warnings.append({
                    "rule": "W-03",
                    "name": "strategic_fit_above_cap",
                    "value": round(sf_val, 2),
                    "threshold": 1.5,
                    "severity": "warning",
                    "message": f"Strategic Fit input {sf_val:.2f}x exceeds 1.5x cap — Priority Score may be inflated"
                })
            else:
                passed_checks.append("strategic_fit_within_cap")
    except Exception:
        pass

    # === W-11: Low source quality — research yielded mostly T4 sources ===
    # This rule reads research_validation.md if present (from Step 2.5)
    try:
        from pathlib import Path as _Path
        # Look for research_validation.md in same directory as workbook
        wb_dir = _Path(workbook_path).parent
        rv_path = wb_dir / 'research_validation.md'
        if not rv_path.exists():
            # Try with feature code prefix
            for f in wb_dir.glob('*research_validation.md'):
                rv_path = f
                break

        if rv_path.exists():
            content = rv_path.read_text(encoding='utf-8', errors='ignore')
            # Count tier mentions in audit
            t2_count = content.count('Tier T2 sources |')
            t3_count = content.count('Tier T3 sources |')
            t4_count = content.count('Tier T4 sources |')
            fallback_count_match = re.search(r'Fallbacks .*?\| (\d+) \|', content)
            fallbacks = int(fallback_count_match.group(1)) if fallback_count_match else 0

            # Heuristic: if total fallbacks ≥ 50% of CFs, low quality
            cf_count_match = re.search(r'Total CFs researched \| (\d+) \|', content)
            cf_count = int(cf_count_match.group(1)) if cf_count_match else 0

            if cf_count > 0 and fallbacks / cf_count >= 0.5:
                warnings.append({
                    "rule": "W-11",
                    "name": "low_source_quality",
                    "value": f"{fallbacks}/{cf_count} CFs fell back",
                    "severity": "warning",
                    "message": (
                        f"Research yielded fallbacks for {fallbacks} of {cf_count} CFs "
                        "(≥50%). ROI numbers rely heavily on library defaults. "
                        "Re-run research or accept lower confidence."
                    )
                })
            elif cf_count > 0:
                passed_checks.append("research_quality_acceptable")
    except Exception:
        # Research file not found or parse error — silently skip
        # (Pre-research workbooks are valid, just no W-11 check)
        pass


    # === FI-01..FI-03: Formula integrity (uses formula-view workbook) ===
    fi_result = _check_formula_integrity(wb_f)
    red_flags.extend(fi_result['red_flags'])
    passed_checks.extend(fi_result['passed_checks'])

    if len(red_flags) >= 1:
        trigger_review = True
        trigger_reason = f"Triggered: {len(red_flags)} red flag(s)"
    elif len(warnings) >= 2:
        trigger_review = True
        trigger_reason = f"Triggered: {len(warnings)} warnings (≥2 threshold)"
    elif user_high_stakes:
        trigger_review = True
        trigger_reason = "Triggered: user requested deep review"
    else:
        trigger_review = False
        if len(warnings) == 1:
            trigger_reason = "Skipped: 1 warning (within threshold), clean otherwise"
        else:
            trigger_reason = "Skipped: clean output, ROI within normal range"

    return {
        "status": "success",
        "workbook": str(workbook_path),
        "red_flags": red_flags,
        "warnings": warnings,
        "passed_checks": passed_checks,
        "trigger_review": trigger_review,
        "trigger_reason": trigger_reason,
        "metrics": {
            "y1_roi_base": round(y1_roi_base, 2) if isinstance(y1_roi_base, (int, float)) else None,
            "y1_roi_best": round(y1_roi_best, 2) if isinstance(y1_roi_best, (int, float)) else None,
            "y3_roi_base": round(y3_roi_base, 2) if isinstance(y3_roi_base, (int, float)) else None,
        }
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_roi.py <workbook.xlsx> [--user-high-stakes]")
        sys.exit(1)

    workbook_path = Path(sys.argv[1])
    user_high_stakes = '--user-high-stakes' in sys.argv

    if not workbook_path.exists():
        print(json.dumps({"status": "error", "message": f"File not found: {workbook_path}"}))
        sys.exit(1)

    result = validate(workbook_path, user_high_stakes=user_high_stakes)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == '__main__':
    main()
