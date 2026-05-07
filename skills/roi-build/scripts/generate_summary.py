#!/usr/bin/env python3
"""
Generate summary.md from ROI workbook + validator output + optional reviews.

Usage:
    python generate_summary.py <workbook.xlsx> <validator.json> <output.md> [cfo_review.md] [hop_review.md]

Reads:
- ROI numbers from xlsx Sheet 5
- Driver tree from Sheet 1
- Risks/validation from Sheet 6
- Validator flags from JSON
- Top 3 from review files (if exist)

Output: filled summary.md
"""

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


def fmt_money(v):
    """Format THB value with commas."""
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        if abs(v) >= 1_000_000:
            return f"{v/1_000_000:,.2f}M"
        return f"{v:,.0f}"
    return str(v)


def fmt_mult(v):
    """Format ROI multiplier."""
    if v is None or not isinstance(v, (int, float)):
        return "—"
    return f"{v:.2f}x"


def extract_review_top3(review_path):
    """Extract Q1/Q2/Q3 or P1/P2/P3 from a review markdown file."""
    if not review_path or not Path(review_path).exists():
        return None

    content = Path(review_path).read_text(encoding='utf-8')
    # Match "### Q1: ..." or "### P1: ..." lines
    pattern = re.compile(r'###\s+([QP][123]):?\s+(.+?)(?=\n###|\n##|\Z)', re.DOTALL)
    matches = pattern.findall(content)

    if not matches:
        return None

    items = []
    for tag, body in matches[:3]:
        # Take first line as headline
        first_line = body.strip().split('\n')[0].strip()
        items.append(f"  {tag}: {first_line[:200]}")
    return "\n".join(items)


def extract_xlsx_data(wb):
    """Extract key metrics, driver tree, risks, validation plan from workbook."""
    data = {
        'feature_code': '', 'feature_name': '', 'date': '', 'owner': '', 'status': '',
        'hypothesis': '', 'primary_outcome': '', 'driver_tree': '',
        'cagan': {'value': '—', 'usability': '—', 'feasibility': '—', 'viability': '—'},
        'roi': {},
        'risks': [],
        'validation': [],
        'flagged': [],
    }

    # Sheet 1: Feature_Info
    try:
        ws = wb['1_Feature_Info']
        # Title row
        title = ws['A1'].value or ''
        if ' — ' in title:
            data['feature_code'], data['feature_name'] = title.split(' — ', 1)
        elif ' - ' in title:
            data['feature_code'], data['feature_name'] = title.split(' - ', 1)

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if not cell.value:
                    continue
                key = str(cell.value).strip()
                next_val = ws.cell(row=cell.row, column=2).value
                if key == 'Feature Code': data['feature_code'] = next_val or data['feature_code']
                elif key == 'Feature Name': data['feature_name'] = next_val or data['feature_name']
                elif key == 'Date': data['date'] = next_val or ''
                elif key == 'Owner': data['owner'] = next_val or ''
                elif key == 'Status': data['status'] = next_val or ''
                elif key == 'Hypothesis': data['hypothesis'] = next_val or ''
                elif key == 'Primary Outcome Metric': data['primary_outcome'] = next_val or ''
                elif key == 'Mechanism (Driver Tree)': data['driver_tree'] = next_val or ''
                elif key == 'Value Risk': data['cagan']['value'] = next_val or '—'
                elif key == 'Usability Risk': data['cagan']['usability'] = next_val or '—'
                elif key == 'Feasibility Risk': data['cagan']['feasibility'] = next_val or '—'
                elif key == 'Viability Risk': data['cagan']['viability'] = next_val or '—'
    except Exception as e:
        print(f"WARN: Sheet 1 parse: {e}", file=sys.stderr)

    # Sheet 5: Output — load with data_only for calculated values
    try:
        ws = wb['5_Output']
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                key = cell.value.strip()
                if key == 'Total Value (THB)':
                    data['roi']['y1_val_w'] = ws.cell(row=cell.row, column=2).value
                    data['roi']['y1_val_b'] = ws.cell(row=cell.row, column=3).value
                    data['roi']['y1_val_best'] = ws.cell(row=cell.row, column=4).value
                elif key == 'Y1 Cost (Build)':
                    data['roi']['y1_cost_w'] = ws.cell(row=cell.row, column=2).value
                    data['roi']['y1_cost_b'] = ws.cell(row=cell.row, column=3).value
                    data['roi']['y1_cost_best'] = ws.cell(row=cell.row, column=4).value
                elif key == 'Y1 ROI (Value ÷ Cost)':
                    data['roi']['y1_roi_w'] = ws.cell(row=cell.row, column=2).value
                    data['roi']['y1_roi_b'] = ws.cell(row=cell.row, column=3).value
                    data['roi']['y1_roi_best'] = ws.cell(row=cell.row, column=4).value
                elif key == '3-Year ROI':
                    data['roi']['y3_roi_w'] = ws.cell(row=cell.row, column=2).value
                    data['roi']['y3_roi_b'] = ws.cell(row=cell.row, column=3).value
                    data['roi']['y3_roi_best'] = ws.cell(row=cell.row, column=4).value
    except Exception as e:
        print(f"WARN: Sheet 5 parse: {e}", file=sys.stderr)

    # Sheet 4: Effort
    try:
        ws = wb['4_Effort_Cost']
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                if cell.value.strip() == 'TOTAL BUILD EFFORT (MD)':
                    data['roi']['effort_w'] = ws.cell(row=cell.row, column=2).value
                    data['roi']['effort_b'] = ws.cell(row=cell.row, column=3).value
                    data['roi']['effort_best'] = ws.cell(row=cell.row, column=4).value
                    break
    except Exception as e:
        print(f"WARN: Sheet 4 parse: {e}", file=sys.stderr)

    # Sheet 6: Risks + Validation
    try:
        ws = wb['6_Flagged_Assumptions']
        capture = None  # 'risks' or 'validation' or 'flagged'
        for row in ws.iter_rows(values_only=False):
            first = row[0].value if row else None
            if isinstance(first, str):
                if 'TOP 3 RISKS' in first.upper():
                    capture = 'risks'
                    continue
                elif 'VALIDATION PLAN' in first.upper():
                    capture = 'validation'
                    continue
                elif 'FLAGGED ASSUMPTIONS' in first.upper():
                    capture = 'flagged'
                    continue
                elif first in ('#', 'Time'):
                    continue
            if capture == 'risks' and isinstance(first, int):
                data['risks'].append({
                    '#': first,
                    'risk': row[1].value or '',
                    'impact': row[2].value or '',
                    'likelihood': row[3].value or '',
                    'mitigation': row[4].value or '',
                })
            elif capture == 'validation' and first and not isinstance(first, int):
                data['validation'].append({
                    'time': first,
                    'indicator': row[1].value or '',
                    'target': row[2].value or '',
                    'kill': row[3].value or '',
                })
            elif capture == 'flagged' and isinstance(first, int):
                data['flagged'].append({
                    '#': first,
                    'assumption': row[1].value or '',
                    'value': row[2].value or '',
                    'tier': row[3].value or '',
                })
    except Exception as e:
        print(f"WARN: Sheet 6 parse: {e}", file=sys.stderr)

    return data


def render_summary(data, validator, cfo_top3=None, hop_top3=None):
    """Render summary.md content."""
    roi = data['roi']
    code = data.get('feature_code', 'N/A')
    name = data.get('feature_name', '')

    # Decision tier from Y1 Base ROI
    y1_roi_b = roi.get('y1_roi_b')
    y3_roi_b = roi.get('y3_roi_b')

    def tier_y1(v):
        if not isinstance(v, (int, float)): return '—'
        if v >= 3: return '🟢 STRONG GO'
        if v >= 1: return '🟡 GO'
        if v >= 0.5: return '🟠 DEFER'
        return '🔴 KILL'

    def tier_y3(v):
        if not isinstance(v, (int, float)): return '—'
        if v >= 5: return '🟢 STRONG GO'
        if v >= 2: return '🟡 GO'
        if v >= 1: return '🟠 MARGINAL'
        return '🔴 KILL'

    lines = []
    lines.append(f"# {code} — {name} ROI Summary")
    lines.append("")
    lines.append(f"**Date:** {data.get('date', '')}")
    lines.append(f"**Owner:** {data.get('owner', '')}")
    lines.append(f"**Status:** {data.get('status', '')}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Bottom Line
    lines.append("## Bottom Line")
    lines.append("")
    lines.append("| Metric | Worst | **Base** | Best |")
    lines.append("|--------|-------|----------|------|")
    lines.append(f"| Y1 Total Value (THB) | {fmt_money(roi.get('y1_val_w'))} | **{fmt_money(roi.get('y1_val_b'))}** | {fmt_money(roi.get('y1_val_best'))} |")
    lines.append(f"| Y1 Cost (Build) | {fmt_money(roi.get('y1_cost_w'))} | **{fmt_money(roi.get('y1_cost_b'))}** | {fmt_money(roi.get('y1_cost_best'))} |")
    lines.append(f"| **Y1 ROI** | {fmt_mult(roi.get('y1_roi_w'))} | **{fmt_mult(roi.get('y1_roi_b'))}** | {fmt_mult(roi.get('y1_roi_best'))} |")
    lines.append(f"| 3-Year ROI | {fmt_mult(roi.get('y3_roi_w'))} | **{fmt_mult(roi.get('y3_roi_b'))}** | {fmt_mult(roi.get('y3_roi_best'))} |")
    lines.append(f"| Build Effort | {roi.get('effort_w','—')} MD | {roi.get('effort_b','—')} MD | {roi.get('effort_best','—')} MD |")
    lines.append("")
    lines.append(f"**Y1 Decision:** {tier_y1(y1_roi_b)}")
    lines.append(f"**3-Year Decision:** {tier_y3(y3_roi_b)}")
    lines.append("")

    # Hypothesis
    if data.get('hypothesis'):
        lines.append("---")
        lines.append("")
        lines.append("## Hypothesis")
        lines.append("")
        lines.append(data['hypothesis'])
        lines.append("")
        if data.get('primary_outcome'):
            lines.append(f"**Primary outcome metric:** {data['primary_outcome']}")
            lines.append("")

    # Driver Tree
    if data.get('driver_tree'):
        lines.append("---")
        lines.append("")
        lines.append("## Driver Tree")
        lines.append("")
        lines.append("```")
        lines.append(str(data['driver_tree']))
        lines.append("```")
        lines.append("")

    # Risks
    if data.get('risks'):
        lines.append("---")
        lines.append("")
        lines.append("## Top Risks")
        lines.append("")
        lines.append("| # | Risk | Impact | Likelihood | Mitigation |")
        lines.append("|---|------|--------|------------|------------|")
        for risk in data['risks']:
            lines.append(f"| {risk['#']} | {risk['risk']} | {risk['impact']} | {risk['likelihood']} | {risk['mitigation']} |")
        lines.append("")

    # Validation Plan
    if data.get('validation'):
        lines.append("---")
        lines.append("")
        lines.append("## Pilot / Validation Plan")
        lines.append("")
        lines.append("| Time | Indicator | Target | Kill |")
        lines.append("|------|-----------|--------|------|")
        for v in data['validation']:
            lines.append(f"| {v['time']} | {v['indicator']} | {v['target']} | {v['kill']} |")
        lines.append("")

    # Cagan 4 Risks
    cg = data['cagan']
    lines.append("---")
    lines.append("")
    lines.append("## Cagan 4-Risk Status")
    lines.append("")
    lines.append("| Risk | Status |")
    lines.append("|------|--------|")
    lines.append(f"| Value | {cg['value']} |")
    lines.append(f"| Usability | {cg['usability']} |")
    lines.append(f"| Feasibility | {cg['feasibility']} |")
    lines.append(f"| Viability | {cg['viability']} |")
    lines.append("")

    # Validator
    lines.append("---")
    lines.append("")
    lines.append("## Validator Output")
    lines.append("")
    rf = validator.get('red_flags', [])
    wn = validator.get('warnings', [])
    pc = validator.get('passed_checks', [])
    lines.append(f"- **Red flags:** {len(rf)}")
    lines.append(f"- **Warnings:** {len(wn)}")
    lines.append(f"- **Passed checks:** {len(pc)}")
    lines.append(f"- **Trigger review:** {validator.get('trigger_review', False)}")
    lines.append(f"- **Reason:** {validator.get('trigger_reason', '')}")
    lines.append("")
    if rf:
        lines.append("### 🔴 Red Flags")
        for f in rf:
            lines.append(f"- **{f.get('rule','?')}**: {f.get('message','')}")
        lines.append("")
    if wn:
        lines.append("### 🟡 Warnings")
        for w in wn:
            lines.append(f"- **{w.get('rule','?')}**: {w.get('message','')}")
        lines.append("")
    if pc:
        lines.append("### ✅ Passed Checks")
        lines.append(f"{', '.join(pc)}")
        lines.append("")

    # Reviewer notes (if any)
    if cfo_top3 or hop_top3:
        lines.append("---")
        lines.append("")
        lines.append("## Reviewer Notes")
        lines.append("")
        if cfo_top3:
            lines.append("### 💼 Skeptical CFO")
            lines.append(cfo_top3)
            lines.append("")
        if hop_top3:
            lines.append("### 🎯 Pragmatic Head of Product (Cagan)")
            lines.append(hop_top3)
            lines.append("")

    # Top assumptions
    if data.get('flagged'):
        lines.append("---")
        lines.append("")
        lines.append("## Top Assumptions to Verify")
        lines.append("")
        for a in data['flagged'][:3]:
            lines.append(f"{a['#']}. **{a['assumption']}** ({a.get('tier','T?')}) — {a.get('value','')}")
        lines.append("")

    return "\n".join(lines)


def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_summary.py <workbook.xlsx> <validator.json> <output.md> [cfo_review.md] [hop_review.md]")
        sys.exit(1)

    workbook_path = Path(sys.argv[1])
    validator_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    cfo_path = sys.argv[4] if len(sys.argv) > 4 else None
    hop_path = sys.argv[5] if len(sys.argv) > 5 else None

    if not workbook_path.exists():
        print(f"Error: workbook not found: {workbook_path}")
        sys.exit(1)

    # Load with data_only=True so we get calculated values, not formulas
    wb = load_workbook(workbook_path, data_only=True)
    data = extract_xlsx_data(wb)

    with open(validator_path) as f:
        validator = json.load(f)

    cfo_top3 = extract_review_top3(cfo_path)
    hop_top3 = extract_review_top3(hop_path)

    summary = render_summary(data, validator, cfo_top3, hop_top3)
    output_path.write_text(summary, encoding='utf-8')

    print(json.dumps({
        "status": "success",
        "output": str(output_path),
        "feature_code": data.get('feature_code', ''),
        "lines": len(summary.split('\n')),
    }))


if __name__ == '__main__':
    main()
