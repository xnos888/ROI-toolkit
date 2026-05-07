#!/usr/bin/env python3
"""
Cache computed formula values into xlsx WITHOUT destroying formulas.

Strategy:
1. Compute all formulas via `formulas` library (in-memory)
2. Open the xlsx as a zip archive
3. For each sheet XML, find each formula cell <c><f>...</f></c>
4. Inject <v>computed_value</v> after </f> — keeping <f> intact
5. Repack the xlsx

Result: xlsx file has BOTH:
  - <f>=A1+B1</f>  (formula — visible/editable in Excel)
  - <v>30.5</v>    (cached value — shown when xlsx opens, used by openpyxl data_only=True)

Usage:
    python recalc.py <workbook.xlsx>
"""
import sys
import re
import shutil
import tempfile
import zipfile
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

import formulas
import openpyxl


def _compute(path: str) -> dict:
    """Compute all formulas via `formulas` lib. Return {SHEET_UPPER: {cell_ref: value}}."""
    xl = formulas.ExcelModel().loads(path).finish()
    sol = xl.calculate()
    fname = Path(path).name.upper()
    cached: dict = {}
    for key, val in sol.items():
        if not isinstance(key, str) or not key.startswith("'["):
            continue
        try:
            head, cell_ref = key.rsplit("'!", 1)
            inside = head.lstrip("'").lstrip("[")
            file_part, sheet = inside.split("]", 1)
            if file_part.upper() != fname:
                continue
            v = val.value
            if hasattr(v, '__iter__') and not isinstance(v, (str, bytes)):
                try:
                    flat = list(v)
                    if flat and hasattr(flat[0], '__iter__') and not isinstance(flat[0], (str, bytes)):
                        v = flat[0][0] if len(flat[0]) > 0 else None
                    elif flat:
                        v = flat[0]
                    else:
                        v = None
                except Exception:
                    continue
            cls = type(v).__name__
            if cls in ('XlError', 'XlErrors'):
                continue
            if isinstance(v, str) and v.startswith('#'):
                continue
            if v is None:
                continue
            cached.setdefault(sheet.upper(), {})[cell_ref] = v
        except Exception:
            continue
    return cached


def _format_value(v) -> str:
    """Format computed value for XML <v> tag."""
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, (int, float)):
        # Avoid scientific notation for normal numbers
        if isinstance(v, float) and abs(v) < 1e15 and v == int(v):
            return str(int(v))
        return repr(v)
    return str(v)


def inject_cached_values(path: str) -> tuple:
    """Inject cached values into xlsx. Returns (cells_updated, cells_with_errors)."""
    cached = _compute(path)
    if not cached:
        return 0, 0

    # Map sheet display name -> xml file path
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    # openpyxl saves sheets as sheet1.xml, sheet2.xml in order of sheetnames
    sheet_to_xml = {}
    for i, name in enumerate(sheet_names, 1):
        sheet_to_xml[name.upper()] = f"xl/worksheets/sheet{i}.xml"

    # Reverse: xml path -> sheet upper name
    xml_to_sheet = {v: k for k, v in sheet_to_xml.items()}

    # Regex to match formula cells:
    # <c r="A1" ...><f...>...</f>[<v />|<v>...</v>]?</c>
    pattern = re.compile(
        r'(<c\b[^>]*\br="([A-Z]+\d+)"[^>]*>)'             # opening <c r="..."> with attrs
        r'(<f\b[^>]*>(?:[^<]|<(?!/f>))*</f>)'             # <f>...</f> formula
        r'(?:<v\s*/>|<v\b[^>]*>[^<]*</v>)?'               # optional empty/filled <v> (discarded)
        r'(</c>)',                                         # closing </c>
        re.DOTALL,
    )

    updated_total = 0
    skipped_total = 0

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        with zipfile.ZipFile(path, 'r') as zin, \
             zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                content = zin.read(item)
                if item in xml_to_sheet:
                    sheet_upper = xml_to_sheet[item]
                    sheet_cache = cached.get(sheet_upper, {})
                    if sheet_cache:
                        text = content.decode('utf-8')
                        local_updated = 0
                        local_skipped = 0

                        def inject(m):
                            nonlocal local_updated, local_skipped
                            open_tag, ref, formula_block, close_tag = m.groups()
                            v = sheet_cache.get(ref)
                            if v is None:
                                local_skipped += 1
                                return m.group(0)
                            # Decide cell type and format value
                            if isinstance(v, bool):
                                clean_open = re.sub(r'\s+t="[^"]*"', '', open_tag)
                                clean_open = clean_open[:-1] + ' t="b">'
                                v_str = '1' if v else '0'
                            elif isinstance(v, (int, float)):
                                clean_open = re.sub(r'\s+t="[^"]*"', '', open_tag)
                                v_str = _format_value(v)
                            else:
                                # String — use t="str" (formula string result)
                                clean_open = re.sub(r'\s+t="[^"]*"', '', open_tag)
                                clean_open = clean_open[:-1] + ' t="str">'
                                # XML-escape string
                                v_str = (str(v).replace('&', '&amp;').replace('<', '&lt;')
                                         .replace('>', '&gt;'))
                            local_updated += 1
                            return f'{clean_open}{formula_block}<v>{v_str}</v>{close_tag}'

                        new_text = pattern.sub(inject, text)
                        content = new_text.encode('utf-8')
                        updated_total += local_updated
                        skipped_total += local_skipped
                zout.writestr(item, content)

        shutil.move(tmp.name, path)
    except Exception:
        Path(tmp.name).unlink(missing_ok=True)
        raise

    return updated_total, skipped_total


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <workbook.xlsx>")
        sys.exit(1)
    path = sys.argv[1]
    updated, skipped = inject_cached_values(path)
    print(f'{{"status": "success", "workbook": "{path}", "cached_values_injected": {updated}, "skipped_no_value": {skipped}}}')


if __name__ == '__main__':
    main()
