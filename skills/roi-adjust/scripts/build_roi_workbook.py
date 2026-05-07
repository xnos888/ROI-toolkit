#!/usr/bin/env python3
"""
Build 6-sheet ROI workbook from JSON inputs.

Usage:
    python build_roi_workbook.py <inputs.json> <output_dir>

Output: <output_dir>/<FEATURE_CODE>_ROI.xlsx

Inputs JSON schema: see templates/input_form.yaml for documented format.
"""

import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============== HELPERS ==============

def resolve_ref(value, baseline_map, context=""):
    """
    Strict resolver for '=ref:metric_name' syntax.

    Returns:
        (resolved_value, kind) where kind in {'literal', 'formula', 'ref'}

    Raises ValueError with clear message if reference not found.
    """
    if not isinstance(value, str):
        return value, "literal"

    # Match =ref:... case-insensitively, allow whitespace
    match = re.match(r'^\s*=\s*ref\s*:\s*(.+?)\s*$', value, re.IGNORECASE)
    if not match:
        # Plain formula or literal string
        if value.startswith('='):
            return value, "formula"
        return value, "literal"

    metric = match.group(1).strip()
    if metric not in baseline_map:
        available = "\n".join(f"   - '{m}'" for m in baseline_map.keys())
        raise ValueError(
            f"❌ Reference '={value}' could not be resolved.\n"
            f"   Context: {context}\n"
            f"   Looking for: '{metric}'\n"
            f"   Available baseline metrics:\n{available}\n"
            f"   Fix: ensure metric name in baseline section matches exactly (case-sensitive)."
        )
    return f"=B{baseline_map[metric]}", "ref"


# ============== STYLES ==============
FONT_BASE = "Arial"
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name=FONT_BASE, bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', start_color='D9E1F2')
SECTION_FONT = Font(name=FONT_BASE, bold=True, color='1F4E78', size=11)
SUB_FILL = PatternFill('solid', start_color='E2EFDA')
SUB_FONT = Font(name=FONT_BASE, bold=True, color='375623', size=10)
TOTAL_FILL = PatternFill('solid', start_color='FFE699')
INPUT_FILL = PatternFill('solid', start_color='FFF2CC')
INPUT_FONT = Font(name=FONT_BASE, color='0000FF', size=10)
CALC_FONT = Font(name=FONT_BASE, color='000000', size=10)
CROSSREF_FONT = Font(name=FONT_BASE, color='006100', size=10)
NOTE_FONT = Font(name=FONT_BASE, italic=True, color='666666', size=9)
BOLD = Font(name=FONT_BASE, bold=True, size=10)
BOLD_LARGE = Font(name=FONT_BASE, bold=True, size=12, color='1F4E78')
BOLD_HUGE = Font(name=FONT_BASE, bold=True, size=14, color='1F4E78')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURR = '#,##0;(#,##0);"-"'
CURR_M = '#,##0.0,,"M";(#,##0.0,,"M");"-"'
PCT = '0.0%;(0.0%);"-"'
MULT = '0.00"x"'
INT = '#,##0;(#,##0);"-"'


def H(c):
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER


def S(c):
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')


def SUB(c):
    c.fill = SUB_FILL
    c.font = SUB_FONT


def IN(c, fmt=None):
    c.fill = INPUT_FILL
    c.font = INPUT_FONT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def CA(c, fmt=None):
    c.font = CALC_FONT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def XR(c, fmt=None):
    c.font = CROSSREF_FONT
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def TOTAL(c, fmt=None):
    c.fill = TOTAL_FILL
    c.font = BOLD
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============== SHEET BUILDERS ==============

def build_sheet1_feature_info(wb, inputs):
    """Sheet 1: Feature_Info — Identity + Problem + Mechanism + Cagan 4 Risks"""
    ws = wb.create_sheet('1_Feature_Info')
    widths(ws, [28, 70])

    code = inputs['feature']['code']
    name = inputs['feature']['name']
    ws['A1'] = f"{code} — {name}"
    ws['A1'].font = Font(name=FONT_BASE, bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:B1')

    # IDENTITY
    r = 3
    ws.cell(row=r, column=1, value='IDENTITY')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    f = inputs['feature']
    identity = [
        ('Feature Code', f['code']),
        ('Feature Name', f['name']),
        ('Category', f.get('category', 'Unspecified')),
        ('Sub-feature Scope', f.get('scope', '')),
        ('Owner', f.get('owner', '')),
        ('Status', f.get('status', 'Discovery → ROI Assessment')),
        ('Date', f.get('date', '')),
        ('Confidence Tier', inputs['confidence'].get('tier', 'T3')),
    ]
    for k, v in identity:
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=2, value=v).font = CALC_FONT
        r += 1

    # PROBLEM & MECHANISM
    r += 1
    ws.cell(row=r, column=1, value='PROBLEM & MECHANISM')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    mechanism = [
        ('Problem (Why)', f.get('problem', '')),
        ('What it does', f.get('what_it_does', '')),
        ('Mechanism (Driver Tree)', inputs.get('driver_tree', '')),
        ('Primary Outcome Metric', inputs.get('primary_outcome', '')),
        ('Secondary Metrics', inputs.get('secondary_metrics', '')),
        ('Hypothesis', inputs.get('hypothesis', '')),
    ]
    for k, v in mechanism:
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
        cv = ws.cell(row=r, column=2, value=v)
        cv.font = CALC_FONT
        cv.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 60 if len(str(v)) > 80 else 20
        r += 1

    # CAGAN 4 RISKS
    r += 1
    ws.cell(row=r, column=1, value="CAGAN'S 4 RISKS — DISCOVERY CHECK")
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    risks = inputs.get('cagan_risks', {
        'Value Risk': 'Not assessed',
        'Usability Risk': 'Not assessed',
        'Feasibility Risk': 'Not assessed',
        'Viability Risk': 'Not assessed',
    })
    for k, v in risks.items():
        ws.cell(row=r, column=1, value=k).font = BOLD
        cv = ws.cell(row=r, column=2, value=v)
        cv.font = CALC_FONT
        cv.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1


def build_sheet2_inputs(wb, inputs):
    """Sheet 2: Inputs — Baseline + TAM-SAM-SOM + CFs + Confidence/SF"""
    ws = wb.create_sheet('2_Inputs')
    widths(ws, [38, 16, 16, 16, 12, 50])

    ws['A1'] = 'INPUTS — Baseline + TAM-SAM-SOM + Conversion Factors'
    ws['A1'].font = BOLD_LARGE
    ws.merge_cells('A1:F1')

    ws['A2'] = 'Yellow = Input (changeable) | Black = formula | Green = cross-sheet ref'
    ws['A2'].font = NOTE_FONT
    ws.merge_cells('A2:F2')

    # === BASELINE ===
    r = 4
    ws.cell(row=r, column=1, value='BASELINE DATA')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['Metric', 'Value', '', '', 'Unit', 'Source / Notes']
    for i, h in enumerate(hdr, 1):
        if h:
            H(ws.cell(row=r, column=i, value=h))
    r += 1

    baseline = inputs.get('baseline', [])
    baseline_row_map = {}  # name → row for cross-ref
    for item in baseline:
        ws.cell(row=r, column=1, value=item['metric']).font = CALC_FONT
        c = ws.cell(row=r, column=2, value=item['value'])
        fmt = PCT if item.get('format') == 'pct' else INT
        IN(c, fmt)
        ws.cell(row=r, column=5, value=item.get('unit', '')).font = NOTE_FONT
        ws.cell(row=r, column=6, value=item.get('source', '')).font = NOTE_FONT
        baseline_row_map[item['metric']] = r
        r += 1

    # === TAM-SAM-SOM ===
    r += 1
    ws.cell(row=r, column=1, value='TAM-SAM-SOM (3 Scenarios)')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['Layer', 'Worst', 'Base', 'Best', 'Unit', 'Logic / Source']
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1

    tss = inputs['tam_sam_som']

    # TAM
    ws.cell(row=r, column=1, value=tss['tam']['label']).font = BOLD
    for col, key in [(2, 'worst'), (3, 'base'), (4, 'best')]:
        v = tss['tam'][key]
        resolved, kind = resolve_ref(v, baseline_row_map,
                                      context=f"TAM {key} ({tss['tam'].get('label','')})")
        if kind == "ref":
            XR(ws.cell(row=r, column=col, value=resolved), INT)
        elif kind == "formula":
            CA(ws.cell(row=r, column=col, value=resolved), INT)
        else:
            IN(ws.cell(row=r, column=col, value=resolved), INT)
    ws.cell(row=r, column=5, value=tss['tam'].get('unit', 'VN/yr')).font = NOTE_FONT
    ws.cell(row=r, column=6, value=tss['tam'].get('source', '')).font = NOTE_FONT
    TAM_ROW = r
    r += 1

    # SAM filters
    sam_rows = []
    for filt in tss.get('sam_filters', []):
        ws.cell(row=r, column=1, value=filt['label']).font = CALC_FONT
        IN(ws.cell(row=r, column=2, value=filt['worst']), PCT)
        IN(ws.cell(row=r, column=3, value=filt['base']), PCT)
        IN(ws.cell(row=r, column=4, value=filt['best']), PCT)
        ws.cell(row=r, column=5, value='%').font = NOTE_FONT
        ws.cell(row=r, column=6, value=filt.get('source', '')).font = NOTE_FONT
        sam_rows.append(r)
        r += 1

    # SAM total
    ws.cell(row=r, column=1, value='SAM: VN reachable per year').font = BOLD
    sam_formula_w = f'=B{TAM_ROW}' + ''.join(f'*B{sr}' for sr in sam_rows)
    sam_formula_b = f'=C{TAM_ROW}' + ''.join(f'*C{sr}' for sr in sam_rows)
    sam_formula_best = f'=D{TAM_ROW}' + ''.join(f'*D{sr}' for sr in sam_rows)
    CA(ws.cell(row=r, column=2, value=sam_formula_w), INT)
    CA(ws.cell(row=r, column=3, value=sam_formula_b), INT)
    CA(ws.cell(row=r, column=4, value=sam_formula_best), INT)
    ws.cell(row=r, column=5, value='VN/yr').font = NOTE_FONT
    ws.cell(row=r, column=6, value='Auto-calc').font = NOTE_FONT
    SAM_ROW = r
    r += 1

    # SOM ramp Y1, Y2, Y3
    som_rows = {}
    for year_label, year_key in [('Y1', 'som_y1'), ('Y2', 'som_y2'), ('Y3', 'som_y3')]:
        som = tss[year_key]
        ws.cell(row=r, column=1, value=f'SOM {year_label} adoption').font = CALC_FONT
        IN(ws.cell(row=r, column=2, value=som['worst']), PCT)
        IN(ws.cell(row=r, column=3, value=som['base']), PCT)
        IN(ws.cell(row=r, column=4, value=som['best']), PCT)
        ws.cell(row=r, column=5, value='%').font = NOTE_FONT
        ws.cell(row=r, column=6, value=som.get('source', '')).font = NOTE_FONT
        som_rows[year_label] = r
        r += 1

    # Effective addressable VN
    r += 1
    ws.cell(row=r, column=1, value='EFFECTIVE ADDRESSABLE VN').font = BOLD
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    eff_rows = {}
    for year_label in ['Y1', 'Y2', 'Y3']:
        ws.cell(row=r, column=1, value=f'{year_label} effective VN').font = BOLD
        sr = som_rows[year_label]
        CA(ws.cell(row=r, column=2, value=f'=B{SAM_ROW}*B{sr}'), INT)
        CA(ws.cell(row=r, column=3, value=f'=C{SAM_ROW}*C{sr}'), INT)
        CA(ws.cell(row=r, column=4, value=f'=D{SAM_ROW}*D{sr}'), INT)
        eff_rows[year_label] = r
        r += 1

    # === CONVERSION FACTORS ===
    r += 2
    ws.cell(row=r, column=1, value='CONVERSION FACTORS')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['Driver', 'Worst', 'Base', 'Best', 'Tier', 'Source + Why']
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1

    cf_rows = {}
    for cf in inputs['conversion_factors']:
        ws.cell(row=r, column=1, value=cf['label']).font = BOLD if cf.get('primary') else CALC_FONT
        v_w = cf['worst']
        v_b = cf['base']
        v_best = cf['best']
        # Detect format
        if cf.get('format') == 'pct':
            fmt = PCT
        elif cf.get('format') == 'currency':
            fmt = INT
        else:
            fmt = '0.00'
        # Allow formula refs (e.g., reference to baseline)
        for col, val in [(2, v_w), (3, v_b), (4, v_best)]:
            resolved, kind = resolve_ref(val, baseline_row_map,
                                          context=f"CF {cf['id']} ({cf['label']})")
            if kind == "ref":
                XR(ws.cell(row=r, column=col, value=resolved), fmt)
            elif kind == "formula":
                CA(ws.cell(row=r, column=col, value=resolved), fmt)
            else:
                IN(ws.cell(row=r, column=col, value=resolved), fmt)
        ws.cell(row=r, column=5, value=cf.get('tier', 'T3')).font = NOTE_FONT
        ws.cell(row=r, column=6, value=cf.get('source', '')).font = NOTE_FONT
        cf_rows[cf['id']] = r
        r += 1

    # === CONFIDENCE & STRATEGIC FIT ===
    r += 1
    ws.cell(row=r, column=1, value='CONFIDENCE & STRATEGIC FIT')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    conf = inputs['confidence']
    ws.cell(row=r, column=1, value='Confidence Tier').font = BOLD
    c = ws.cell(row=r, column=3, value=conf.get('tier', 'T3'))
    c.font = BOLD
    c.alignment = Alignment(horizontal='center')
    c.border = BORDER
    ws.cell(row=r, column=6, value=f"{conf.get('reason', '')} (metadata only — not multiplied into value)").font = NOTE_FONT
    CONF_TIER_ROW = r
    r += 1

    sf = inputs['strategic_fit']
    ws.cell(row=r, column=1, value='Strategic Fit (Priority Score Input)').font = BOLD
    IN(ws.cell(row=r, column=3, value=sf['multiplier']), '0.00')
    ws.cell(row=r, column=6, value=f"{sf.get('reason', '')} (used for Priority Score — never multiplied into Pure ROI)").font = NOTE_FONT
    SF_ROW = r
    r += 1

    return {
        'TAM_ROW': TAM_ROW,
        'SAM_ROW': SAM_ROW,
        'EFF_Y1': eff_rows['Y1'],
        'EFF_Y2': eff_rows['Y2'],
        'EFF_Y3': eff_rows['Y3'],
        'CF_ROWS': cf_rows,
        'CONF_TIER_ROW': CONF_TIER_ROW,
        'SF_ROW': SF_ROW,
        'BASELINE_ROW_MAP': baseline_row_map,
    }


def build_sheet3_value_calc(wb, inputs, refs):
    """Sheet 3: Value_Calc — Y1/Y2/Y3 + 3-Year Cumulative"""
    ws = wb.create_sheet('3_Value_Calc')
    widths(ws, [42, 16, 16, 16, 6, 50])

    ws['A1'] = 'VALUE CALCULATION — 3 Scenarios × 3 Years'
    ws['A1'].font = BOLD_LARGE
    ws.merge_cells('A1:F1')

    INP = "'2_Inputs'!"
    SF = f"{INP}C{refs['SF_ROW']}"  # used by Sheet 5 Priority Score only

    r = 3
    # Driver tree summary
    ws.cell(row=r, column=1, value='DRIVER TREE')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(row=r, column=1, value=inputs.get('driver_tree', '')).font = NOTE_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 80
    r += 2

    components = inputs['value_components']  # list of components defining the calc

    year_adj_rows = {}
    for year_label, year_key in [('YEAR 1', 'Y1'), ('YEAR 2', 'Y2'), ('YEAR 3', 'Y3')]:
        ws.cell(row=r, column=1, value=year_label)
        S(ws.cell(row=r, column=1))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

        hdr = ['Component', 'Worst', 'Base', 'Best', '', 'Formula']
        for i, h in enumerate(hdr, 1):
            if h:
                H(ws.cell(row=r, column=i, value=h))
        r += 1

        # Effective addressable
        eff_row_in_inputs = refs[f'EFF_{year_key}']
        ws.cell(row=r, column=1, value='Effective Addressable VN').font = BOLD
        XR(ws.cell(row=r, column=2, value=f"={INP}B{eff_row_in_inputs}"), INT)
        XR(ws.cell(row=r, column=3, value=f"={INP}C{eff_row_in_inputs}"), INT)
        XR(ws.cell(row=r, column=4, value=f"={INP}D{eff_row_in_inputs}"), INT)
        ws.cell(row=r, column=6, value='From Inputs SAM × SOM').font = NOTE_FONT
        EFF_ROW_LOCAL = r
        r += 1

        # Build value components
        component_total_rows = []
        for comp in components:
            # Section header
            ws.cell(row=r, column=1, value=comp['name']).font = BOLD
            SUB(ws.cell(row=r, column=1))
            r += 1

            # Calculate via formula chain
            prev_row = EFF_ROW_LOCAL
            for step in comp['steps']:
                ws.cell(row=r, column=1, value=f"  {step['label']}").font = CALC_FONT
                # Build formula based on step type
                if step['type'] == 'cohort':
                    # Pass-through: cohort VN is already in EFF_ROW_LOCAL.
                    # Do not advance prev_row; skip writing formula (label-only row).
                    ws.cell(row=r, column=6, value=step.get('note', '')).font = NOTE_FONT
                    r += 1
                    continue
                elif step['type'] == 'fixed_cf_product':
                    # Product of named CFs independent of cohort VN.
                    # step['cf_ids'] = list of CF ids to multiply together.
                    cf_ids = step.get('cf_ids', [])
                    cf_rows = [refs['CF_ROWS'].get(cid) for cid in cf_ids if refs['CF_ROWS'].get(cid)]
                    if cf_rows:
                        for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
                            parts = [f"{INP}{col_letter}{cr}" for cr in cf_rows]
                            formula = '=' + '*'.join(parts)
                            CA(ws.cell(row=r, column=col, value=formula),
                               CURR if step.get('output_currency') else INT)
                elif step['type'] == 'multiply_cf':
                    cf_id = step['cf_id']
                    cf_row = refs['CF_ROWS'].get(cf_id)
                    if cf_row:
                        for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
                            formula = f"={col_letter}{prev_row}*{INP}{col_letter}{cf_row}"
                            CA(ws.cell(row=r, column=col, value=formula),
                               INT if not step.get('output_currency') else CURR)
                elif step['type'] == 'multiply_baseline':
                    baseline_metric = step['metric']
                    bl_row = refs['BASELINE_ROW_MAP'].get(baseline_metric)
                    if bl_row:
                        for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
                            formula = f"={col_letter}{prev_row}*{INP}B{bl_row}"
                            CA(ws.cell(row=r, column=col, value=formula),
                               CURR if step.get('output_currency') else INT)
                elif step['type'] == 'multiply_static':
                    factor = step['factor']
                    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
                        formula = f"={col_letter}{prev_row}*{factor}"
                        CA(ws.cell(row=r, column=col, value=formula),
                           CURR if step.get('output_currency') else INT)

                ws.cell(row=r, column=6, value=step.get('note', '')).font = NOTE_FONT
                prev_row = r
                r += 1

            component_total_rows.append(prev_row)

        # Total Value (sum of components — Pure, no multipliers)
        r += 1
        ws.cell(row=r, column=1, value=f'{year_label} TOTAL VALUE').font = BOLD
        for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
            f = '=' + '+'.join(f"{col_letter}{ctr}" for ctr in component_total_rows)
            TOTAL(ws.cell(row=r, column=col, value=f), CURR)
        ws.cell(row=r, column=6, value='Pure (sum of components, no multipliers)').font = NOTE_FONT
        year_adj_rows[year_key] = r  # key name kept for backward compat in callers
        r += 2

    # 3-Year Cumulative
    ws.cell(row=r, column=1, value='3-YEAR CUMULATIVE')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['', 'Worst', 'Base', 'Best', '', '']
    for i, h in enumerate(hdr, 1):
        if h:
            H(ws.cell(row=r, column=i, value=h))
    r += 1
    ws.cell(row=r, column=1, value='3-Year Total Value').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        f = f'={col_letter}{year_adj_rows["Y1"]}+{col_letter}{year_adj_rows["Y2"]}+{col_letter}{year_adj_rows["Y3"]}'
        TOTAL(ws.cell(row=r, column=col, value=f), CURR)
    CUM_ROW = r

    return {
        'Y1_ADJ_ROW': year_adj_rows['Y1'],
        'Y2_ADJ_ROW': year_adj_rows['Y2'],
        'Y3_ADJ_ROW': year_adj_rows['Y3'],
        'CUM_ROW': CUM_ROW,
    }


def build_sheet4_effort_cost(wb, inputs):
    """Sheet 4: Effort_Cost"""
    ws = wb.create_sheet('4_Effort_Cost')
    widths(ws, [32, 14, 14, 14, 14, 40])

    ws['A1'] = 'EFFORT & COST'
    ws['A1'].font = BOLD_LARGE
    ws.merge_cells('A1:F1')

    r = 3
    ws.cell(row=r, column=1, value='ASSUMPTIONS')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    md_cost = inputs.get('effort', {}).get('md_cost', 24000)
    ma_rate = inputs.get('effort', {}).get('ma_rate', 0.30)

    ws.cell(row=r, column=1, value='Blended MD Cost (THB/MD)').font = BOLD
    IN(ws.cell(row=r, column=2, value=md_cost), INT)
    ws.cell(row=r, column=6, value='Blended SWE/PM/PD/QA').font = NOTE_FONT
    MD_COST_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='MA Rate (%/year of build cost)').font = BOLD
    IN(ws.cell(row=r, column=2, value=ma_rate), PCT)
    ws.cell(row=r, column=6, value='Default 30% (20% × 1.5 multi-branch)').font = NOTE_FONT
    MA_RATE_ROW = r
    r += 2

    # Build effort
    ws.cell(row=r, column=1, value='BUILD EFFORT (MD by Role × Scenario)')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['Role / Task', 'Worst (MD)', 'Base (MD)', 'Best (MD)', '', 'Notes']
    for i, h in enumerate(hdr, 1):
        if h:
            H(ws.cell(row=r, column=i, value=h))
    r += 1

    effort_rows = []
    for item in inputs['effort']['breakdown']:
        ws.cell(row=r, column=1, value=item['role']).font = CALC_FONT
        IN(ws.cell(row=r, column=2, value=item['worst']))
        IN(ws.cell(row=r, column=3, value=item['base']))
        IN(ws.cell(row=r, column=4, value=item['best']))
        ws.cell(row=r, column=6, value=item.get('note', '')).font = NOTE_FONT
        effort_rows.append(r)
        r += 1

    r += 1
    start_r = effort_rows[0]
    end_r = effort_rows[-1]
    ws.cell(row=r, column=1, value='TOTAL BUILD EFFORT (MD)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        TOTAL(ws.cell(row=r, column=col, value=f'=SUM({col_letter}{start_r}:{col_letter}{end_r})'), INT)
    TOT_MD_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='TOTAL BUILD COST (THB)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        TOTAL(ws.cell(row=r, column=col, value=f'={col_letter}{TOT_MD_ROW}*B{MD_COST_ROW}'), CURR)
    BUILD_COST_ROW = r
    r += 2

    ws.cell(row=r, column=1, value='ANNUAL MA COST (THB/yr from Y2)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        TOTAL(ws.cell(row=r, column=col, value=f'={col_letter}{BUILD_COST_ROW}*B{MA_RATE_ROW}'), CURR)
    MA_COST_ROW = r
    r += 2

    ws.cell(row=r, column=1, value='Year 1 Total Cost (Build only)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        CA(ws.cell(row=r, column=col, value=f'={col_letter}{BUILD_COST_ROW}'), CURR)
    Y1_COST_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='3-Year Total Cost (Build + 2 yrs MA)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        CA(ws.cell(row=r, column=col, value=f'={col_letter}{BUILD_COST_ROW}+{col_letter}{MA_COST_ROW}*2'), CURR)
    COST_3YR_ROW = r

    return {
        'TOT_MD_ROW': TOT_MD_ROW,
        'BUILD_COST_ROW': BUILD_COST_ROW,
        'MA_COST_ROW': MA_COST_ROW,
        'Y1_COST_ROW': Y1_COST_ROW,
        'COST_3YR_ROW': COST_3YR_ROW,
    }


def build_sheet5_output(wb, inputs, vc_refs, ec_refs, inp_refs):
    """Sheet 5: Output — ROI Decision Summary (Pure ROI + Priority Score)"""
    ws = wb.create_sheet('5_Output')
    widths(ws, [32, 18, 18, 18, 50])

    ws['A1'] = 'OUTPUT — ROI Decision Summary'
    ws['A1'].font = BOLD_HUGE
    ws.merge_cells('A1:E1')

    VC = "'3_Value_Calc'!"
    EC = "'4_Effort_Cost'!"
    INP = "'2_Inputs'!"
    SF_REF = f"{INP}C{inp_refs['SF_ROW']}"
    Y1_VAL_W = f"{VC}B{vc_refs['Y1_ADJ_ROW']}"
    Y1_VAL_B = f"{VC}C{vc_refs['Y1_ADJ_ROW']}"
    Y1_VAL_BEST = f"{VC}D{vc_refs['Y1_ADJ_ROW']}"
    CUM_VAL_W = f"{VC}B{vc_refs['CUM_ROW']}"
    CUM_VAL_B = f"{VC}C{vc_refs['CUM_ROW']}"
    CUM_VAL_BEST = f"{VC}D{vc_refs['CUM_ROW']}"
    Y1_COST_W = f"{EC}B{ec_refs['Y1_COST_ROW']}"
    Y1_COST_B = f"{EC}C{ec_refs['Y1_COST_ROW']}"
    Y1_COST_BEST = f"{EC}D{ec_refs['Y1_COST_ROW']}"
    Y3_COST_W = f"{EC}B{ec_refs['COST_3YR_ROW']}"
    Y3_COST_B = f"{EC}C{ec_refs['COST_3YR_ROW']}"
    Y3_COST_BEST = f"{EC}D{ec_refs['COST_3YR_ROW']}"

    # Y1 Summary
    r = 3
    ws.cell(row=r, column=1, value='YEAR 1 SUMMARY')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    hdr = ['Metric', 'Worst', 'Base', 'Best', 'Notes']
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1

    ws.cell(row=r, column=1, value='Total Value (THB)').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'={Y1_VAL_W}'), CURR_M)
    XR(ws.cell(row=r, column=3, value=f'={Y1_VAL_B}'), CURR_M)
    XR(ws.cell(row=r, column=4, value=f'={Y1_VAL_BEST}'), CURR_M)
    ws.cell(row=r, column=5, value='Pure value — sum of components, no multipliers').font = NOTE_FONT
    Y1_VAL_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='Y1 Cost (Build)').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'={Y1_COST_W}'), CURR_M)
    XR(ws.cell(row=r, column=3, value=f'={Y1_COST_B}'), CURR_M)
    XR(ws.cell(row=r, column=4, value=f'={Y1_COST_BEST}'), CURR_M)
    Y1_COST_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='Y1 Net Value').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        CA(ws.cell(row=r, column=col, value=f'={col_letter}{Y1_VAL_ROW}-{col_letter}{Y1_COST_ROW}'), CURR_M)
    r += 1

    ws.cell(row=r, column=1, value='Y1 ROI (Value ÷ Cost)').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        TOTAL(ws.cell(row=r, column=col, value=f'=IFERROR({col_letter}{Y1_VAL_ROW}/{col_letter}{Y1_COST_ROW},0)'), MULT)
    ws.cell(row=r, column=5, value='≥1x payback Y1, ≥3x strong').font = NOTE_FONT
    Y1_ROI_ROW = r
    r += 2

    # 3-Year Summary
    ws.cell(row=r, column=1, value='3-YEAR CUMULATIVE SUMMARY')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1

    ws.cell(row=r, column=1, value='3-Year Cumulative Value').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'={CUM_VAL_W}'), CURR_M)
    XR(ws.cell(row=r, column=3, value=f'={CUM_VAL_B}'), CURR_M)
    XR(ws.cell(row=r, column=4, value=f'={CUM_VAL_BEST}'), CURR_M)
    Y3_VAL_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='3-Year Cumulative Cost').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'={Y3_COST_W}'), CURR_M)
    XR(ws.cell(row=r, column=3, value=f'={Y3_COST_B}'), CURR_M)
    XR(ws.cell(row=r, column=4, value=f'={Y3_COST_BEST}'), CURR_M)
    Y3_COST_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='3-Year Net Value').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        CA(ws.cell(row=r, column=col, value=f'={col_letter}{Y3_VAL_ROW}-{col_letter}{Y3_COST_ROW}'), CURR_M)
    r += 1

    ws.cell(row=r, column=1, value='3-Year ROI').font = BOLD
    for col, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
        TOTAL(ws.cell(row=r, column=col, value=f'=IFERROR({col_letter}{Y3_VAL_ROW}/{col_letter}{Y3_COST_ROW},0)'), MULT)
    Y3_ROI_ROW = r
    r += 2

    # Decision
    ws.cell(row=r, column=1, value='DECISION')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    ws.cell(row=r, column=1, value='Y1 Base ROI').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'=C{Y1_ROI_ROW}'), MULT)
    ws.cell(row=r, column=3, value='Tier:').font = BOLD
    ws.cell(row=r, column=4, value=f'=IF(C{Y1_ROI_ROW}>=3,"🟢 STRONG GO",IF(C{Y1_ROI_ROW}>=1,"🟡 GO",IF(C{Y1_ROI_ROW}>=0.5,"🟠 DEFER","🔴 KILL")))').font = BOLD
    r += 1

    ws.cell(row=r, column=1, value='3-Year Base ROI').font = BOLD
    XR(ws.cell(row=r, column=2, value=f'=C{Y3_ROI_ROW}'), MULT)
    ws.cell(row=r, column=3, value='Tier:').font = BOLD
    ws.cell(row=r, column=4, value=f'=IF(C{Y3_ROI_ROW}>=5,"🟢 STRONG GO",IF(C{Y3_ROI_ROW}>=2,"🟡 GO",IF(C{Y3_ROI_ROW}>=1,"🟠 MARGINAL","🔴 KILL")))').font = BOLD
    r += 2

    # Priority Score (roadmap ranking only — NOT a value claim)
    ws.cell(row=r, column=1, value='PRIORITY SCORE (roadmap ranking only — not a value claim)')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    ws.cell(row=r, column=1, value='Y1 Priority Score (Base ROI × Strategic Fit)').font = BOLD
    TOTAL(ws.cell(row=r, column=3, value=f'=C{Y1_ROI_ROW}*{SF_REF}'), MULT)
    ws.cell(row=r, column=5, value='Use only for roadmap sequencing, not as headline ROI').font = NOTE_FONT
    PRIORITY_Y1_ROW = r
    r += 1

    ws.cell(row=r, column=1, value='3-Yr Priority Score (3-Yr Base ROI × Strategic Fit)').font = BOLD
    TOTAL(ws.cell(row=r, column=3, value=f'=C{Y3_ROI_ROW}*{SF_REF}'), MULT)
    PRIORITY_Y3_ROW = r
    r += 2

    # CEO 1-liner
    ws.cell(row=r, column=1, value='CEO 1-LINER').font = BOLD
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

    code = inputs['feature']['code']
    name = inputs['feature']['name']
    ws.cell(row=r, column=1, value=(
        f'=CONCATENATE("{code} {name} (",TEXT(\'4_Effort_Cost\'!C{ec_refs["TOT_MD_ROW"]},"#,##0")," MD): '
        f'Y1 ",TEXT(C{Y1_ROI_ROW},"0.00")," x | 3-yr ",TEXT(C{Y3_ROI_ROW},"0.00")," x | Priority ",TEXT(C{PRIORITY_Y1_ROW},"0.00")," x")'
    )).font = BOLD
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    # Post-launch tracking
    ws.cell(row=r, column=1, value='POST-LAUNCH TRACKING (fill after launch)')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    hdr = ['Metric', 'Predicted (Base)', 'Actual', 'Variance %', 'Notes']
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1
    for tracking in inputs.get('post_launch_tracking', []):
        ws.cell(row=r, column=1, value=tracking['metric']).font = CALC_FONT
        ws.cell(row=r, column=2, value=tracking['predicted']).font = NOTE_FONT
        IN(ws.cell(row=r, column=3))
        ws.cell(row=r, column=5, value=tracking.get('source', '')).font = NOTE_FONT
        r += 1

    return {
        'Y1_ROI_ROW': Y1_ROI_ROW,
        'Y3_ROI_ROW': Y3_ROI_ROW,
    }


def build_sheet6_flagged(wb, inputs):
    """Sheet 6: Flagged_Assumptions + Risk Register + Validation Plan"""
    ws = wb.create_sheet('6_Flagged_Assumptions')
    widths(ws, [6, 38, 18, 12, 30, 30])

    ws['A1'] = 'FLAGGED ASSUMPTIONS — Verify Before Commit'
    ws['A1'].font = BOLD_LARGE
    ws.merge_cells('A1:F1')

    r = 3
    hdr = ['#', 'Assumption', 'Value (Base)', 'Tier', 'Source', 'Verification Action']
    for i, h in enumerate(hdr, 1):
        H(ws.cell(row=r, column=i, value=h))
    r += 1

    for i, a in enumerate(inputs.get('flagged_assumptions', []), 1):
        for col, val in enumerate([i, a['assumption'], a['value'], a['tier'], a.get('source', ''), a.get('verification', '')], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = CALC_FONT if col != 4 else BOLD
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1

    # Risk register
    r += 2
    ws.cell(row=r, column=1, value='TOP 3 RISKS THAT COULD KILL ROI')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['#', 'Risk', 'Impact', 'Likelihood', 'Mitigation', '']
    for i, h in enumerate(hdr, 1):
        if h:
            H(ws.cell(row=r, column=i, value=h))
    r += 1
    for i, risk in enumerate(inputs.get('risks', []), 1):
        for col, val in enumerate([i, risk['risk'], risk['impact'], risk['likelihood'], risk['mitigation']], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = CALC_FONT
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = BORDER
        ws.row_dimensions[r].height = 35
        r += 1

    # Validation plan
    r += 2
    ws.cell(row=r, column=1, value='VALIDATION PLAN — Leading Indicators')
    S(ws.cell(row=r, column=1))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    hdr = ['Time', 'Leading Indicator', 'Target (Base)', 'Kill Threshold', '', '']
    for i, h in enumerate(hdr, 1):
        if h:
            H(ws.cell(row=r, column=i, value=h))
    r += 1
    for v in inputs.get('validation_plan', []):
        for col, val in enumerate([v['time'], v['indicator'], v['target'], v['kill']], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = CALC_FONT
            c.border = BORDER
        r += 1


# ============== POST-FIX ==============

def fix_accidental_formulas(wb):
    """Notes starting with '=' are mistakenly treated as formulas. Fix them."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    v = cell.value
                    # Detect if it's a note (descriptive text, not a formula)
                    has_descriptive = any(c in v for c in ['×', 'cost/', '(refill', '(incremental',
                                                            '(avg', 'rate)', 'CF-', 'Auto-calc',
                                                            'min)', 'visit)', 'event)'])
                    has_formula_op = any(c in v for c in ['+', '*', '/', 'SUM', 'IF', 'IFERROR',
                                                           'CONCATENATE', 'TEXT', '$', '!'])
                    # If looks like note, strip leading =
                    if has_descriptive and not has_formula_op:
                        cell.value = v[1:].lstrip()


# ============== MAIN ==============

def main():
    if len(sys.argv) < 3:
        print("Usage: python build_roi_workbook.py <inputs.json> <output_dir>")
        sys.exit(1)

    inputs_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)

    with open(inputs_path) as f:
        inputs = json.load(f)

    feature_code = inputs['feature']['code']
    output_path = output_dir / f"{feature_code.replace(' ', '_').replace('/', '_')}_ROI.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    build_sheet1_feature_info(wb, inputs)
    refs = build_sheet2_inputs(wb, inputs)
    vc_refs = build_sheet3_value_calc(wb, inputs, refs)
    ec_refs = build_sheet4_effort_cost(wb, inputs)
    build_sheet5_output(wb, inputs, vc_refs, ec_refs, refs)
    build_sheet6_flagged(wb, inputs)

    fix_accidental_formulas(wb)

    wb.save(output_path)
    print(json.dumps({
        "status": "success",
        "output_path": str(output_path),
        "feature_code": feature_code,
        "sheets_count": 6,
    }))


if __name__ == '__main__':
    main()
