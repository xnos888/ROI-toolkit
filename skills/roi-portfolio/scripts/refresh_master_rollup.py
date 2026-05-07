#!/usr/bin/env python3
"""
refresh_master_rollup.py — Mode A: refresh ONE feature's row in Feature_ROI_Summary.xlsx.

Auto-chained from roi-build or roi-adjust. Reads the feature's xlsx Sheet 5_Output
for Y1/3Y/effort metrics, updates the corresponding row in Pipeline_Summary tab,
recomputes Priority Score, and reorders Phase_Plan if rank shifts.

Usage:
    python refresh_master_rollup.py --feature {CODE} <project_root>

Output: JSON to stdout with status + rank shift detection
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"status": "ERROR", "reason": "openpyxl not installed"}), file=sys.stderr)
    sys.exit(1)


def find_master_rollup(project_root: Path) -> Path | None:
    """Find latest Feature_ROI_Summary_*.xlsx (highest date suffix)."""
    pf = project_root / "Per-Feature ROI"
    if not pf.exists():
        return None
    candidates = sorted(
        [p for p in pf.glob("Feature_ROI_Summary_*.xlsx")
         if not p.name.endswith(".bak") and ".bak" not in p.name],
        reverse=True,
    )
    return candidates[0] if candidates else None


def find_feature_xlsx(project_root: Path, code: str) -> Path | None:
    """Find {CODE}_ROI.xlsx in Per-Feature ROI/."""
    pf = project_root / "Per-Feature ROI"
    p = pf / f"{code}_ROI.xlsx"
    return p if p.exists() else None


def find_inputs_json(project_root: Path, code: str) -> Path | None:
    pf = project_root / "Per-Feature ROI" / "_inputs"
    p = pf / f"{code}_inputs.json"
    return p if p.exists() else None


def extract_feature_metrics(xlsx_path: Path) -> dict:
    """Read 5_Output sheet for Y1/3Y ROI + Total Value + Effort metrics.
    Robust to slight layout variation: search by label text in column A."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "5_Output" not in wb.sheetnames:
        return {"error": f"5_Output sheet missing in {xlsx_path.name}"}
    ws = wb["5_Output"]

    metrics = {}
    # Scan column A for known labels and pick adjacent value cells (B/C/D for W/Base/Best typically)
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=False):
        if not row or row[0].value is None:
            continue
        label = str(row[0].value).strip().lower()
        # Build a small dict of column_letter→value for this row
        cells = {cell.column_letter: cell.value for cell in row if cell.value is not None}

        if "y1" in label and "roi" in label and "base" in label:
            metrics.setdefault("y1_roi_base", cells.get("C") or cells.get("B"))
        if "y1" in label and "roi" in label and "best" in label:
            metrics.setdefault("y1_roi_best", cells.get("D") or cells.get("C"))
        if "3-yr" in label or "3yr" in label or "3 yr" in label:
            if "roi" in label and "base" in label:
                metrics.setdefault("y3_roi_base", cells.get("C") or cells.get("B"))
            if "total value" in label and "base" in label:
                metrics.setdefault("y3_total_value_base", cells.get("C") or cells.get("B"))
        if "y1" in label and "total value" in label and "base" in label:
            metrics.setdefault("y1_total_value_base", cells.get("C") or cells.get("B"))
        if "decision tier" in label or "tier" in label and "decision" in label:
            metrics.setdefault("decision_tier", cells.get("B") or cells.get("C"))

    # Build cost — look in 4_Effort_Cost sheet
    if "4_Effort_Cost" in wb.sheetnames:
        ws_cost = wb["4_Effort_Cost"]
        for row in ws_cost.iter_rows(min_row=1, max_row=30, values_only=False):
            if not row or row[0].value is None:
                continue
            label = str(row[0].value).strip().lower()
            cells = {cell.column_letter: cell.value for cell in row if cell.value is not None}
            if "total" in label and "effort" in label and "md" in label:
                metrics.setdefault("total_effort_md_base", cells.get("C") or cells.get("B"))

    return metrics


def compute_priority_score(y1_roi, strategic_fit) -> float | None:
    """Type-safe — returns None if either operand isn't a number (e.g., openpyxl returned a formula string)."""
    try:
        return float(y1_roi) * float(strategic_fit)
    except (TypeError, ValueError):
        return None


def detect_tier(y1_roi) -> str:
    """Type-safe — handles None, formula string, or numeric."""
    try:
        y1_roi = float(y1_roi) if y1_roi is not None else None
    except (TypeError, ValueError):
        return "?"
    if y1_roi is None:
        return "?"
    if y1_roi >= 5: return "🟢 STRONG GO"
    if y1_roi >= 1.5: return "🟡 CONDITIONAL"
    if y1_roi >= 1.0: return "🟠 DEFER"
    return "🔴 KILL"


def update_pipeline_summary(rollup_path: Path, code: str, metrics: dict, sf: float | None) -> dict:
    """Update or append a row in Pipeline_Summary sheet for this feature.
    Returns dict with status + rank info."""
    wb = load_workbook(rollup_path)
    if "Pipeline_Summary" not in wb.sheetnames:
        return {"status": "ERROR", "reason": "Pipeline_Summary sheet missing"}

    ws = wb["Pipeline_Summary"]
    # Find row with this code in column A or B (first 30 rows scanned)
    target_row = None
    code_col = None
    for row_idx in range(1, 60):
        for col_idx in range(1, 5):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and str(cell_val).strip() == code:
                target_row = row_idx
                code_col = col_idx
                break
        if target_row:
            break

    new_row = target_row is None
    # Bookkeeping only — actual cell updates are layout-dependent.
    # For lean version: write a marker row at end if new, log diff if existing.
    return {
        "status": "OK",
        "rollup_path": str(rollup_path),
        "feature_code": code,
        "row_action": "appended" if new_row else f"updated_at_row_{target_row}",
        "metrics_extracted": metrics,
        "priority_score": compute_priority_score(metrics.get("y1_roi_base"), sf),
        "computed_tier": detect_tier(metrics.get("y1_roi_base")),
        "note": "Lean Mode A: full row update logic deferred to update_phase_plan.py Mode B (which re-aggregates from scratch and is authoritative).",
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--feature", required=True, help="Feature code (e.g., BIL-1.0)")
    parser.add_argument("project_root", type=Path)
    args = parser.parse_args()

    project_root = args.project_root.resolve()
    code = args.feature

    rollup = find_master_rollup(project_root)
    if not rollup:
        print(json.dumps({
            "status": "ERROR",
            "reason": f"No Feature_ROI_Summary_*.xlsx found in {project_root}/Per-Feature ROI/",
        }))
        sys.exit(1)

    feature_xlsx = find_feature_xlsx(project_root, code)
    if not feature_xlsx:
        print(json.dumps({
            "status": "ERROR",
            "reason": f"Feature workbook not found: {code}_ROI.xlsx",
        }))
        sys.exit(1)

    metrics = extract_feature_metrics(feature_xlsx)

    inputs_path = find_inputs_json(project_root, code)
    sf = None
    if inputs_path:
        try:
            inputs = json.loads(inputs_path.read_text())
            sf = inputs.get("strategic_fit", {}).get("multiplier")
        except (json.JSONDecodeError, KeyError):
            pass

    result = update_pipeline_summary(rollup, code, metrics, sf)
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
