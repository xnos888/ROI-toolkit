#!/usr/bin/env python3
"""
Build a single-sheet xlsx comparing N features side-by-side (one column per feature).

Uses Base case values only. Pulls from each feature's inputs.json + validate.json.

Usage:
    python build_comparison.py <output_dir> <feature_code1> <feature_code2> ...

Output: <output_dir>/Batch_Comparison_<date>.xlsx
"""
import sys
import json
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============== STYLES ==============
FONT_BASE = "Arial"
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name=FONT_BASE, bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', start_color='305496')
SECTION_FONT = Font(name=FONT_BASE, bold=True, color='FFFFFF', size=10)
SUB_FILL = PatternFill('solid', start_color='D9E1F2')
SUB_FONT = Font(name=FONT_BASE, bold=True, color='1F4E78', size=10)
TOTAL_FILL = PatternFill('solid', start_color='FFE699')
INPUT_FILL = PatternFill('solid', start_color='FFF2CC')
NOTE_FONT = Font(name=FONT_BASE, italic=True, color='666666', size=9)
BOLD = Font(name=FONT_BASE, bold=True, size=10)
CALC_FONT = Font(name=FONT_BASE, color='000000', size=10)
HIGHLIGHT_FONT = Font(name=FONT_BASE, bold=True, size=11, color='1F4E78')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURR = '#,##0;(#,##0);"-"'
CURR_M = '#,##0.0,,"M";(#,##0.0,,"M");"-"'
PCT = '0.0%;(0.0%);"-"'
MULT = '0.00"x"'
INT = '#,##0;(#,##0);"-"'


def H(c):
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER


def S(c):
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = BORDER


def SUB(c):
    c.fill = SUB_FILL
    c.font = SUB_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')


def TOTAL(c, fmt=None):
    c.fill = TOTAL_FILL
    c.font = BOLD
    c.border = BORDER
    if fmt:
        c.number_format = fmt


def CA(c, fmt=None):
    c.font = CALC_FONT
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(vertical='top', wrap_text=True)


def write_section(ws, r, label, ncols):
    cell = ws.cell(row=r, column=1, value=label)
    S(cell)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    return r + 1


def write_subsection(ws, r, label, ncols):
    cell = ws.cell(row=r, column=1, value=label)
    SUB(cell)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    return r + 1


def get_base(d, key='base'):
    """Get base value from a dict that has worst/base/best keys."""
    if isinstance(d, dict):
        return d.get(key)
    return d


def load_features(out_dir: Path, codes: list) -> list:
    """Load all (inputs, validate, summary_dict) tuples."""
    features = []
    for code in codes:
        ipath = out_dir / "_inputs" / f"{code}_inputs.json"
        vpath = out_dir / "_inputs" / f"{code}_validate.json"
        if not ipath.exists():
            print(f"⚠️  Missing inputs: {code}", file=sys.stderr)
            continue
        inputs = json.loads(ipath.read_text())
        validate = {}
        if vpath.exists():
            try:
                validate = json.loads(vpath.read_text())
            except Exception:
                pass
        features.append({'code': code, 'inputs': inputs, 'validate': validate})
    return features


def collect_all_cf_ids(features):
    """Get superset of CF IDs across all features (preserve insertion order)."""
    seen = []
    for f in features:
        for cf in f['inputs'].get('conversion_factors', []):
            if cf['id'] not in seen:
                seen.append(cf['id'])
    return seen


def collect_all_roles(features):
    seen = []
    for f in features:
        for r in f['inputs'].get('effort', {}).get('breakdown', []):
            role = r['role']
            if role not in seen:
                seen.append(role)
    return seen


def build(out_dir: Path, codes: list, output_path: Path):
    features = load_features(out_dir, codes)
    if not features:
        print("No features loaded — aborting", file=sys.stderr)
        sys.exit(1)

    n = len(features)
    ncols = 1 + n  # col A = label, col B-N = features

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('Feature_Comparison_Base')

    # Column widths
    ws.column_dimensions['A'].width = 38
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = 28

    # Title
    title = f"PE Feature Comparison — Base Case (n={n})"
    ws.cell(row=1, column=1, value=title).font = Font(name=FONT_BASE, bold=True, size=14, color='1F4E78')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()} | Source: Per-Feature ROI/_inputs/*.json | Mode: Base scenario only").font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    r = 4
    # Header row — feature codes
    H(ws.cell(row=r, column=1, value='Attribute'))
    for i, f in enumerate(features):
        H(ws.cell(row=r, column=2 + i, value=f['code']))
    r += 1

    ws.cell(row=r, column=1, value='Feature Name').font = BOLD
    for i, f in enumerate(features):
        c = ws.cell(row=r, column=2 + i, value=f['inputs']['feature']['name'])
        CA(c)
    r += 2

    # === IDENTITY ===
    r = write_section(ws, r, 'IDENTITY', ncols)
    identity_keys = [
        ('Feature Code', lambda f: f['inputs']['feature']['code']),
        ('Category', lambda f: f['inputs']['feature'].get('category', '')),
        ('Owner', lambda f: f['inputs']['feature'].get('owner', '')),
        ('Status', lambda f: f['inputs']['feature'].get('status', '')),
        ('Date', lambda f: f['inputs']['feature'].get('date', '')),
        ('Confidence Tier', lambda f: f['inputs'].get('confidence', {}).get('tier', '')),
    ]
    for label, fn in identity_keys:
        ws.cell(row=r, column=1, value=label).font = BOLD
        for i, f in enumerate(features):
            CA(ws.cell(row=r, column=2 + i, value=fn(f)))
        r += 1
    r += 1

    # === PROBLEM & MECHANISM ===
    r = write_section(ws, r, 'PROBLEM & MECHANISM', ncols)
    pm_keys = [
        ('Problem', lambda f: f['inputs']['feature'].get('problem', '')),
        ('What it does', lambda f: f['inputs']['feature'].get('what_it_does', '')),
        ('Sub-features', lambda f: '\n'.join(f['inputs']['feature'].get('sub_features', []))),
        ('Driver Tree', lambda f: f['inputs'].get('driver_tree', '')),
        ('Primary Outcome', lambda f: f['inputs'].get('primary_outcome', '')),
        ('Secondary Metrics', lambda f: f['inputs'].get('secondary_metrics', '')),
        ('Hypothesis', lambda f: f['inputs'].get('hypothesis', '')),
    ]
    for label, fn in pm_keys:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
        max_len = 0
        for i, f in enumerate(features):
            v = fn(f)
            CA(ws.cell(row=r, column=2 + i, value=v))
            max_len = max(max_len, len(str(v)))
        ws.row_dimensions[r].height = min(180, max(20, max_len // 30 * 15))
        r += 1
    r += 1

    # === CAGAN 4 RISKS ===
    r = write_section(ws, r, "CAGAN'S 4 RISKS", ncols)
    risk_keys = ['Value Risk', 'Usability Risk', 'Feasibility Risk', 'Viability Risk']
    for risk in risk_keys:
        ws.cell(row=r, column=1, value=risk).font = BOLD
        for i, f in enumerate(features):
            v = f['inputs'].get('cagan_risks', {}).get(risk, '')
            CA(ws.cell(row=r, column=2 + i, value=v))
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1

    # === BASELINE DATA (per-feature, key metrics) ===
    r = write_section(ws, r, 'BASELINE DATA (per-feature)', ncols)
    ws.cell(row=r, column=1, value='Each feature lists its own baseline metrics').font = NOTE_FONT
    r += 1
    for i, f in enumerate(features):
        baseline_text = '\n'.join(
            f"• {b['metric']}: {b['value']:,} {b.get('unit','')}" if isinstance(b.get('value'), (int, float))
            else f"• {b['metric']}: {b['value']} {b.get('unit','')}"
            for b in f['inputs'].get('baseline', [])
        )
        # Skip — done below as a single multi-line cell per feature
        pass

    ws.cell(row=r, column=1, value='Baseline metrics (multi-line)').font = BOLD
    ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
    for i, f in enumerate(features):
        baseline_text = '\n'.join(
            f"• {b['metric']}: {b['value']:,} {b.get('unit','')}" if isinstance(b.get('value'), (int, float))
            else f"• {b['metric']}: {b['value']} {b.get('unit','')}"
            for b in f['inputs'].get('baseline', [])
        )
        CA(ws.cell(row=r, column=2 + i, value=baseline_text))
    ws.row_dimensions[r].height = 140
    r += 2

    # === TAM-SAM-SOM (Base) ===
    r = write_section(ws, r, 'TAM-SAM-SOM — Base Scenario', ncols)
    # TAM
    ws.cell(row=r, column=1, value='TAM (Base)').font = BOLD
    for i, f in enumerate(features):
        tam = f['inputs']['tam_sam_som']['tam']
        v = tam.get('base')
        # Resolve =ref if needed
        if isinstance(v, str) and v.startswith('=ref:'):
            metric = v.replace('=ref:', '').strip()
            for b in f['inputs'].get('baseline', []):
                if b['metric'] == metric:
                    v = b['value']
                    break
        unit = tam.get('unit', '')
        c = ws.cell(row=r, column=2 + i, value=v if isinstance(v, (int, float)) else str(v))
        CA(c, INT if isinstance(v, (int, float)) else None)
    r += 1

    # TAM label
    ws.cell(row=r, column=1, value='TAM Description').font = BOLD
    for i, f in enumerate(features):
        CA(ws.cell(row=r, column=2 + i, value=f['inputs']['tam_sam_som']['tam'].get('label', '')))
    ws.row_dimensions[r].height = 40
    r += 1

    # SAM filters
    max_filters = max(len(f['inputs']['tam_sam_som'].get('sam_filters', [])) for f in features)
    for fidx in range(max_filters):
        ws.cell(row=r, column=1, value=f'SAM Filter #{fidx+1} (Base)').font = BOLD
        for i, f in enumerate(features):
            filters = f['inputs']['tam_sam_som'].get('sam_filters', [])
            if fidx < len(filters):
                filt = filters[fidx]
                v = filt.get('base')
                lbl = filt.get('label', '')
                txt = f"{lbl}: {v:.0%}" if isinstance(v, (int, float)) else f"{lbl}: {v}"
                CA(ws.cell(row=r, column=2 + i, value=txt))
            else:
                CA(ws.cell(row=r, column=2 + i, value='—'))
        ws.row_dimensions[r].height = 40
        r += 1

    # SOM Y1 Base
    ws.cell(row=r, column=1, value='SOM Y1 (Base)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs']['tam_sam_som']['som_y1'].get('base')
        c = ws.cell(row=r, column=2 + i, value=v)
        CA(c, PCT)
    r += 1
    ws.cell(row=r, column=1, value='SOM Y2 (Base)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs']['tam_sam_som']['som_y2'].get('base')
        c = ws.cell(row=r, column=2 + i, value=v)
        CA(c, PCT)
    r += 1
    ws.cell(row=r, column=1, value='SOM Y3 (Base)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs']['tam_sam_som']['som_y3'].get('base')
        c = ws.cell(row=r, column=2 + i, value=v)
        CA(c, PCT)
    r += 2

    # === CONVERSION FACTORS (Base) ===
    r = write_section(ws, r, 'CONVERSION FACTORS — Base Scenario', ncols)
    cf_ids = collect_all_cf_ids(features)
    for cf_id in cf_ids:
        # Find label across features (first one with this CF)
        first_label = ''
        for f in features:
            for cf in f['inputs'].get('conversion_factors', []):
                if cf['id'] == cf_id:
                    first_label = cf['label']
                    break
            if first_label:
                break
        ws.cell(row=r, column=1, value=f"{cf_id}: {first_label[:60]}").font = BOLD
        ws.cell(row=r, column=1).alignment = Alignment(vertical='top', wrap_text=True)
        for i, f in enumerate(features):
            this_cf = next((cf for cf in f['inputs'].get('conversion_factors', []) if cf['id'] == cf_id), None)
            if not this_cf:
                CA(ws.cell(row=r, column=2 + i, value='—'))
                continue
            v = this_cf.get('base')
            if isinstance(v, str) and v.startswith('=ref:'):
                metric = v.replace('=ref:', '').strip()
                for b in f['inputs'].get('baseline', []):
                    if b['metric'] == metric:
                        v = b['value']
                        break
            tier = this_cf.get('tier', '')
            fmt_type = this_cf.get('format', 'pct')
            if isinstance(v, (int, float)):
                if fmt_type == 'pct':
                    txt = f"{v:.1%} [{tier}]"
                elif fmt_type == 'currency':
                    txt = f"{v:,.0f} THB [{tier}]"
                else:
                    txt = f"{v:,.2f} [{tier}]"
            else:
                txt = f"{v} [{tier}]"
            CA(ws.cell(row=r, column=2 + i, value=txt))
        ws.row_dimensions[r].height = 24
        r += 1
    r += 1

    # === CONFIDENCE TIER (metadata) + STRATEGIC FIT (Priority Score input) ===
    r = write_section(ws, r, 'CONFIDENCE TIER (metadata) + STRATEGIC FIT (Priority Score input)', ncols)
    ws.cell(row=r, column=1, value='Confidence Tier').font = BOLD
    for i, f in enumerate(features):
        CA(ws.cell(row=r, column=2 + i, value=f['inputs'].get('confidence', {}).get('tier', '')))
    r += 1
    ws.cell(row=r, column=1, value='Confidence Reason').font = BOLD
    for i, f in enumerate(features):
        CA(ws.cell(row=r, column=2 + i, value=f['inputs'].get('confidence', {}).get('reason', '')))
    ws.row_dimensions[r].height = 40
    r += 1
    ws.cell(row=r, column=1, value='Strategic Fit (Priority Score weight)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs'].get('strategic_fit', {}).get('multiplier')
        c = ws.cell(row=r, column=2 + i, value=v)
        CA(c, MULT)
    r += 1
    ws.cell(row=r, column=1, value='Strategic Fit Reason').font = BOLD
    for i, f in enumerate(features):
        CA(ws.cell(row=r, column=2 + i, value=f['inputs'].get('strategic_fit', {}).get('reason', '')))
    ws.row_dimensions[r].height = 40
    r += 2

    # === EFFORT (Base by role) ===
    r = write_section(ws, r, 'EFFORT — Base MD by Role', ncols)
    roles = collect_all_roles(features)
    for role in roles:
        ws.cell(row=r, column=1, value=role[:50]).font = BOLD
        for i, f in enumerate(features):
            breakdown = f['inputs'].get('effort', {}).get('breakdown', [])
            this_role = next((b for b in breakdown if b['role'] == role), None)
            if this_role:
                v = this_role.get('base', 0)
                CA(ws.cell(row=r, column=2 + i, value=v), INT)
            else:
                CA(ws.cell(row=r, column=2 + i, value='—'))
        r += 1
    # Total
    ws.cell(row=r, column=1, value='TOTAL Base MD').font = BOLD
    for i, f in enumerate(features):
        breakdown = f['inputs'].get('effort', {}).get('breakdown', [])
        total = sum(b.get('base', 0) for b in breakdown)
        TOTAL(ws.cell(row=r, column=2 + i, value=total), INT)
    r += 1
    ws.cell(row=r, column=1, value='MD Cost (THB/MD)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs'].get('effort', {}).get('md_cost', 24000)
        CA(ws.cell(row=r, column=2 + i, value=v), CURR)
    r += 1
    ws.cell(row=r, column=1, value='MA Rate (% Y1 ongoing)').font = BOLD
    for i, f in enumerate(features):
        v = f['inputs'].get('effort', {}).get('ma_rate', 0.30)
        CA(ws.cell(row=r, column=2 + i, value=v), PCT)
    r += 2

    # === ROI OUTPUT ===
    r = write_section(ws, r, 'ROI OUTPUT — Base Scenario', ncols)
    ws.cell(row=r, column=1, value='Y1 Base ROI').font = HIGHLIGHT_FONT
    for i, f in enumerate(features):
        v = f.get('validate', {}).get('metrics', {}).get('y1_roi_base')
        c = ws.cell(row=r, column=2 + i, value=v)
        TOTAL(c, MULT)
    r += 1
    ws.cell(row=r, column=1, value='Y1 Best ROI').font = BOLD
    for i, f in enumerate(features):
        v = f.get('validate', {}).get('metrics', {}).get('y1_roi_best')
        c = ws.cell(row=r, column=2 + i, value=v)
        CA(c, MULT)
    r += 1
    ws.cell(row=r, column=1, value='3-Year Base ROI').font = HIGHLIGHT_FONT
    for i, f in enumerate(features):
        v = f.get('validate', {}).get('metrics', {}).get('y3_roi_base')
        c = ws.cell(row=r, column=2 + i, value=v)
        TOTAL(c, MULT)
    r += 1
    # Decision tier (simple based on Y1 base)
    ws.cell(row=r, column=1, value='Decision Tier (Y1 Base)').font = BOLD
    for i, f in enumerate(features):
        roi = f.get('validate', {}).get('metrics', {}).get('y1_roi_base') or 0
        if roi >= 3:
            tier = '🟢 STRONG GO'
        elif roi >= 1:
            tier = '🟡 GO (marginal)'
        elif roi >= 0.5:
            tier = '🟠 NEEDS REFINEMENT'
        else:
            tier = '🔴 KILL'
        CA(ws.cell(row=r, column=2 + i, value=tier))
    r += 2

    # === VALIDATOR FLAGS ===
    r = write_section(ws, r, 'VALIDATOR FLAGS', ncols)
    ws.cell(row=r, column=1, value='Red Flags').font = BOLD
    for i, f in enumerate(features):
        flags = f.get('validate', {}).get('red_flags', [])
        txt = '\n'.join(f"• {fl['rule']}: {fl.get('message','')[:60]}" for fl in flags) or '—'
        CA(ws.cell(row=r, column=2 + i, value=txt))
    ws.row_dimensions[r].height = 60
    r += 1
    ws.cell(row=r, column=1, value='Warnings').font = BOLD
    for i, f in enumerate(features):
        warns = f.get('validate', {}).get('warnings', [])
        txt = '\n'.join(f"• {w['rule']}: {w.get('message','')[:60]}" for w in warns) or '—'
        CA(ws.cell(row=r, column=2 + i, value=txt))
    ws.row_dimensions[r].height = 40
    r += 1
    ws.cell(row=r, column=1, value='Trigger Reason').font = BOLD
    for i, f in enumerate(features):
        CA(ws.cell(row=r, column=2 + i, value=f.get('validate', {}).get('trigger_reason', '')))
    r += 2

    # === RISKS (top from each feature) ===
    r = write_section(ws, r, 'KEY RISKS', ncols)
    ws.cell(row=r, column=1, value='Top Risks (impact/likelihood/mitigation)').font = BOLD
    ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
    for i, f in enumerate(features):
        risks = f['inputs'].get('risks', [])
        txt = '\n\n'.join(
            f"• {rk['risk']}\n  [{rk.get('impact','')}/{rk.get('likelihood','')}] → {rk.get('mitigation','')}"
            for rk in risks[:4]
        )
        CA(ws.cell(row=r, column=2 + i, value=txt))
    ws.row_dimensions[r].height = 220
    r += 2

    # === VALIDATION PLAN ===
    r = write_section(ws, r, 'VALIDATION PLAN — Kill Thresholds', ncols)
    ws.cell(row=r, column=1, value='Kill checkpoints').font = BOLD
    ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
    for i, f in enumerate(features):
        plan = f['inputs'].get('validation_plan', [])
        txt = '\n'.join(
            f"• {p.get('time','')}: {p.get('indicator','')} → target {p.get('target','')}, kill {p.get('kill','')}"
            for p in plan
        )
        CA(ws.cell(row=r, column=2 + i, value=txt))
    ws.row_dimensions[r].height = 140
    r += 2

    # === SOURCE CITATION (per CF Base — abbreviated) ===
    r = write_section(ws, r, 'CF SOURCES (abbreviated)', ncols)
    for cf_id in cf_ids:
        ws.cell(row=r, column=1, value=f"{cf_id} source").font = BOLD
        ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
        for i, f in enumerate(features):
            this_cf = next((cf for cf in f['inputs'].get('conversion_factors', []) if cf['id'] == cf_id), None)
            txt = (this_cf.get('source', '')[:120] + '…') if this_cf and len(this_cf.get('source', '')) > 120 else (this_cf.get('source', '') if this_cf else '—')
            CA(ws.cell(row=r, column=2 + i, value=txt))
        ws.row_dimensions[r].height = 50
        r += 1

    # Freeze pane
    ws.freeze_panes = 'B5'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(json.dumps({
        "status": "success",
        "output": str(output_path),
        "feature_count": n,
        "rows": r,
    }, ensure_ascii=False))


def main():
    if len(sys.argv) < 3:
        print("Usage: python build_comparison.py <output_dir> <feature_code1> [<feature_code2> ...]")
        sys.exit(1)
    out_dir = Path(sys.argv[1])
    codes = sys.argv[2:]
    output_path = out_dir / f"Batch_Comparison_{date.today().isoformat()}.xlsx"
    build(out_dir, codes, output_path)


if __name__ == '__main__':
    main()
