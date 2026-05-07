#!/usr/bin/env python3
"""
update_phase_plan.py — Mode B + C: full batch aggregation + Phase Plan generation + KILL/DEFER proposals.

Mode B: scan all per-feature workbooks, re-aggregate Pipeline_Summary, generate
Phase_Plan tab + markdown narrative, run capacity check per quarter.

Mode C (--propose-kill-defer): rank-order KILL/DEFER candidates per criterion:
    KILL = (Pure ROI < 1.0x) AND (Strategic Fit < 1.2)
    DEFER = (Pure ROI 1.0-1.5x) AND not strategic (SF < 1.4)

Usage:
    python update_phase_plan.py <project_root>
    python update_phase_plan.py --propose-kill-defer <project_root>
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"status": "ERROR", "reason": "openpyxl not installed"}), file=sys.stderr)
    sys.exit(1)


# Reuse extract logic (intentional duplication for skill self-containment)
def extract_feature_metrics(xlsx_path: Path) -> dict:
    """Same as refresh_master_rollup.py — kept here for self-containment."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "5_Output" not in wb.sheetnames:
        return {}
    ws = wb["5_Output"]

    metrics = {}
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=False):
        if not row or row[0].value is None:
            continue
        label = str(row[0].value).strip().lower()
        cells = {cell.column_letter: cell.value for cell in row if cell.value is not None}
        if "y1" in label and "roi" in label and "base" in label:
            metrics.setdefault("y1_roi_base", cells.get("C") or cells.get("B"))
        if "y1" in label and "roi" in label and "best" in label:
            metrics.setdefault("y1_roi_best", cells.get("D") or cells.get("C"))
        if ("3-yr" in label or "3yr" in label) and "roi" in label and "base" in label:
            metrics.setdefault("y3_roi_base", cells.get("C") or cells.get("B"))

    if "4_Effort_Cost" in wb.sheetnames:
        ws_cost = wb["4_Effort_Cost"]
        for row in ws_cost.iter_rows(min_row=1, max_row=30, values_only=False):
            if not row or row[0].value is None:
                continue
            label = str(row[0].value).strip().lower()
            cells = {cell.column_letter: cell.value for cell in row if cell.value is not None}
            if "total" in label and "effort" in label and "md" in label:
                metrics.setdefault("total_effort_md_base", cells.get("C") or cells.get("B"))
    return metrics


def detect_tier(y1_roi):
    if y1_roi is None: return "?"
    try:
        y1_roi = float(y1_roi)
    except (TypeError, ValueError):
        return "?"
    if y1_roi >= 5: return "🟢 STRONG GO"
    if y1_roi >= 1.5: return "🟡 CONDITIONAL"
    if y1_roi >= 1.0: return "🟠 DEFER"
    return "🔴 KILL"


def load_capacity_config(project_root: Path) -> dict:
    """Read _inputs/_capacity_config.json or return defaults."""
    config_path = project_root / "Per-Feature ROI" / "_inputs" / "_capacity_config.json"
    default = {"annual_capacity_md": 600, "buffer_pct": 0.15,
               "quarters": {"Q1": 150, "Q2": 150, "Q3": 150, "Q4": 150}}
    if config_path.exists():
        try:
            return json.loads(config_path.read_text())
        except json.JSONDecodeError:
            return default
    return default


def aggregate_features(project_root: Path) -> list:
    """Scan all {CODE}_ROI.xlsx workbooks, extract metrics + Strategic Fit per feature."""
    pf = project_root / "Per-Feature ROI"
    inputs_dir = pf / "_inputs"
    rows = []

    for xlsx in sorted(pf.glob("*_ROI.xlsx")):
        if ".bak" in xlsx.name or "Comparison" in xlsx.name:
            continue
        code = xlsx.stem.replace("_ROI", "")
        metrics = extract_feature_metrics(xlsx)

        sf = None
        status = None
        inputs_path = inputs_dir / f"{code}_inputs.json"
        if inputs_path.exists():
            try:
                inp = json.loads(inputs_path.read_text())
                sf = inp.get("strategic_fit", {}).get("multiplier")
                status = inp.get("feature", {}).get("status")
            except (json.JSONDecodeError, KeyError):
                pass

        y1 = metrics.get("y1_roi_base")
        try:
            y1_f = float(y1) if y1 is not None else None
        except (TypeError, ValueError):
            y1_f = None

        try:
            sf_f = float(sf) if sf is not None else None
        except (TypeError, ValueError):
            sf_f = None

        priority_score = (y1_f * sf_f) if (y1_f and sf_f) else y1_f

        rows.append({
            "code": code,
            "y1_roi_base": y1_f,
            "y3_roi_base": metrics.get("y3_roi_base"),
            "effort_md_base": metrics.get("total_effort_md_base"),
            "strategic_fit": sf_f,
            "priority_score": priority_score,
            "tier": detect_tier(y1_f),
            "status": status,
        })

    rows.sort(key=lambda r: r["priority_score"] if r["priority_score"] is not None else -1, reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


def assign_quarters(rows: list, capacity: dict) -> list:
    """Greedy quarter assignment by rank, respecting per-quarter capacity."""
    used = {q: 0 for q in capacity["quarters"]}
    quarters_order = list(capacity["quarters"].keys())
    for r in rows:
        if r["status"] in ("KILLED", "DEFERRED"):
            r["quarter"] = "OUT"
            continue
        effort = r.get("effort_md_base") or 0
        try:
            effort = float(effort)
        except (TypeError, ValueError):
            effort = 0
        assigned = False
        for q in quarters_order:
            if used[q] + effort <= capacity["quarters"][q]:
                r["quarter"] = q
                used[q] += effort
                assigned = True
                break
        if not assigned:
            r["quarter"] = "OVERFLOW"
    return rows, used


def propose_kill_defer(rows: list) -> list:
    """Identify KILL/DEFER candidates per Kim's criterion."""
    proposals = []
    for r in rows:
        y1 = r.get("y1_roi_base")
        sf = r.get("strategic_fit") or 1.0
        if y1 is None:
            continue
        if y1 < 1.0 and sf < 1.2:
            proposals.append({**r, "proposed_action": "KILL",
                              "rationale": f"Y1 ROI {y1:.2f}x + SF {sf:.2f} → no financial OR strategic case"})
        elif y1 < 1.5 and sf < 1.4:
            proposals.append({**r, "proposed_action": "DEFER",
                              "rationale": f"Y1 ROI {y1:.2f}x + SF {sf:.2f} → defer until ROI proven via pilot"})
    return proposals


def render_phase_plan_md(rows: list, used: dict, capacity: dict, project_root: Path) -> Path:
    today = date.today().isoformat()
    out_path = project_root / "Per-Feature ROI" / f"PE_Roadmap_Phase_Plan_{today}.md"

    md = f"# PE Roadmap Phase Plan — {today}\n\n"
    md += f"**Annual capacity:** {capacity['annual_capacity_md']} MD ({capacity.get('buffer_pct', 0.15)*100:.0f}% buffer)\n\n"

    md += "## Capacity Utilization\n\n| Quarter | Used | Capacity | Status |\n|---|---|---|---|\n"
    for q, cap in capacity["quarters"].items():
        u = used.get(q, 0)
        flag = "🔴 OVER" if u > cap else "🟡 TIGHT" if u > cap * 0.9 else "🟢 OK"
        md += f"| {q} | {u:.0f} | {cap} | {flag} |\n"

    md += "\n## Feature Ranking (by Priority Score)\n\n"
    md += "| Rank | Code | Tier | Y1 ROI | 3Y ROI | SF | Priority Score | Effort MD | Quarter | Status |\n"
    md += "|---|---|---|---|---|---|---|---|---|---|\n"
    for r in rows:
        y1 = r.get("y1_roi_base") or 0
        y3 = r.get("y3_roi_base") or 0
        sf = r.get("strategic_fit") or 1.0
        ps = r.get("priority_score") or 0
        eff = r.get("effort_md_base") or 0
        try:
            eff_str = f"{float(eff):.0f}"
        except (TypeError, ValueError):
            eff_str = str(eff)
        md += f"| {r['rank']} | {r['code']} | {r['tier']} | {float(y1):.2f}x | {y3} | {sf:.2f} | {ps:.2f} | {eff_str} | {r.get('quarter','?')} | {r.get('status') or '-'} |\n"

    out_path.write_text(md, encoding="utf-8")
    return out_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("project_root", type=Path)
    parser.add_argument("--propose-kill-defer", action="store_true",
                        help="Mode C: propose KILL/DEFER candidates")
    args = parser.parse_args()

    project_root = args.project_root.resolve()
    capacity = load_capacity_config(project_root)
    rows = aggregate_features(project_root)
    rows, used = assign_quarters(rows, capacity)

    result = {
        "status": "OK",
        "n_features": len(rows),
        "capacity_used": used,
        "capacity_total": capacity["quarters"],
        "tier_distribution": {
            "STRONG_GO": sum(1 for r in rows if "STRONG GO" in (r.get("tier") or "")),
            "CONDITIONAL": sum(1 for r in rows if "CONDITIONAL" in (r.get("tier") or "")),
            "DEFER": sum(1 for r in rows if "DEFER" in (r.get("tier") or "")),
            "KILL": sum(1 for r in rows if "KILL" in (r.get("tier") or "")),
        },
        "oversubscription": {q: max(0, used.get(q, 0) - cap) for q, cap in capacity["quarters"].items()},
    }

    if args.propose_kill_defer:
        result["proposals"] = propose_kill_defer(rows)
    else:
        phase_plan_path = render_phase_plan_md(rows, used, capacity, project_root)
        result["phase_plan_md"] = str(phase_plan_path)

    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
